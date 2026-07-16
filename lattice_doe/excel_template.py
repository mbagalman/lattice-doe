# excel_template.py
# License: MIT
"""
Excel workbook bidirectional connector for lattice-doe
=============================================================

Creates a structured ``.xlsx`` workbook pre-filled with a runnable example
and, after a design run, writes Results / Design / Buckets sheets back into
the **same** file.

Sheet layout (Config sheet)
---------------------------
The ``Config`` sheet uses the same sentinel-header structure as the Google
Sheets connector (``sheets.py``), so the same mental model applies:

  ``[SETTINGS]``  — key/value pairs (formula, power parameters, design options).
  ``[CONTRAST]``  — optional; required when ``power_mode = contrast``.
  ``[FACTORS]``   — factor table: Name | Type | Value 1 | Value 2 | …

Dropdown validation is applied to ``power_mode`` (r2 / contrast) and
``criterion`` (I / D / A) cells.

Output sheets
-------------
``excel_run()`` writes three sheets into the workbook after a successful run:

  ``Results``  — summary key/value table (n, power, elapsed, …)
  ``Design``   — full design DataFrame
  ``Buckets``  — run-frequency buckets

Install
-------
  pip install "lattice-doe[extras]"
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

from .config import (
    PowerContrastConfig, PowerR2Config, PowerGLMContrastConfig,
    DesignOptions, MultiResponseOptions, ResponseSpec,
)
from ._request_builder import build_power_cfg, build_design_opts
from .utils import safe_name_slug

# ---------------------------------------------------------------------------
# Soft dependency guard
# ---------------------------------------------------------------------------
try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_INSTALL_HINT = 'pip install "lattice-doe[extras]"'

# ---------------------------------------------------------------------------
# Section sentinels (must match sheets.py constants for consistency)
# ---------------------------------------------------------------------------
_SENTINEL_SETTINGS  = "[SETTINGS]"
_SENTINEL_CONTRAST  = "[CONTRAST]"
_SENTINEL_FACTORS   = "[FACTORS]"
_SENTINEL_RESPONSES = "[RESPONSES]"

# ---------------------------------------------------------------------------
# Custom exception
# ---------------------------------------------------------------------------
class ExcelError(RuntimeError):
    """Raised for all Excel integration failures.

    Covers missing sections, malformed data, parse errors, and write failures.
    The underlying exception is attached as ``__cause__``.
    """


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise ImportError(
            f"openpyxl is required for Excel integration. {_INSTALL_HINT}"
        )


def _sentinel_style(ws: Any, row: int, label: str) -> None:
    """Apply bold dark-blue styling to a sentinel header row."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="left")


def _key_cell(ws: Any, row: int, key: str, value: Any = "") -> None:
    """Write a key in column A (bold) and value in column B."""
    kc = ws.cell(row=row, column=1, value=key)
    kc.font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def _header_row(ws: Any, row: int, headers: List[str]) -> None:
    """Write column headers in bold with a light-blue background."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")


def _set_column_widths(ws: Any, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_dropdown(ws: Any, formula: str, row: int, col: int = 2) -> None:
    """Attach an in-cell dropdown DataValidation to the given cell."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=False,
                        showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=col))


# ---------------------------------------------------------------------------
# Config sheet reader (openpyxl-based, mirrors sheets.py._parse_config_sheet)
# ---------------------------------------------------------------------------
def _read_config_sheet(
    ws: Any,
) -> Tuple[str, Dict[str, Any], Optional[Union[PowerContrastConfig, PowerR2Config]], DesignOptions, Optional[MultiResponseOptions]]:
    """Parse the Config worksheet and return ``(formula, factors, power_cfg, design_opts, multi_cfg)``.

    Raises
    ------
    ExcelError
        For missing sentinels, missing required keys, or malformed data.
    """
    # Collect all rows as lists of stripped strings
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])

    # Locate sentinels
    idx_settings:  Optional[int] = None
    idx_contrast:  Optional[int] = None
    idx_factors:   Optional[int] = None
    idx_responses: Optional[int] = None

    for i, row in enumerate(rows):
        col_a = row[0] if row else ""
        if col_a == _SENTINEL_SETTINGS:
            idx_settings = i
        elif col_a == _SENTINEL_CONTRAST:
            idx_contrast = i
        elif col_a == _SENTINEL_FACTORS:
            idx_factors = i
        elif col_a == _SENTINEL_RESPONSES:
            idx_responses = i

    if idx_settings is None:
        raise ExcelError(
            f"Config sheet is missing the '{_SENTINEL_SETTINGS}' sentinel. "
            "Ensure cell A1 (or equivalent) contains '[SETTINGS]'."
        )
    if idx_factors is None:
        raise ExcelError(
            f"Config sheet is missing the '{_SENTINEL_FACTORS}' sentinel."
        )

    # --- Parse [SETTINGS] ---
    sentinels_after = sorted(
        j for j in (idx_contrast, idx_factors, idx_responses)
        if j is not None and j > idx_settings
    )
    settings_end = sentinels_after[0] if sentinels_after else len(rows)

    settings: Dict[str, str] = {}
    for row in rows[idx_settings + 1: settings_end]:
        col_a = row[0] if len(row) > 0 else ""
        col_b = row[1] if len(row) > 1 else ""
        if col_a:
            settings[col_a] = col_b

    if not settings.get("formula"):
        raise ExcelError(
            "Config sheet [SETTINGS] is missing the required 'formula' key."
        )
    # power_mode is required unless a [RESPONSES] section is present
    if idx_responses is None and not settings.get("power_mode"):
        raise ExcelError(
            "Config sheet [SETTINGS] is missing the required 'power_mode' key."
        )

    formula = settings["formula"]
    power_mode = settings.get("power_mode", "").strip().lower()

    if power_mode and power_mode not in ("r2", "contrast", "glm"):
        raise ExcelError(
            f"[SETTINGS] power_mode must be 'r2', 'contrast', or 'glm'; got {settings.get('power_mode')!r}."
        )

    def _float(key: str, default: float) -> float:
        raw = settings.get(key, "").strip()
        if not raw:
            return default
        try:
            return float(raw)
        except ValueError:
            raise ExcelError(
                f"[SETTINGS] key '{key}' must be a number; got {raw!r}."
            ) from None

    def _int(key: str, default: int) -> int:
        raw = settings.get(key, "").strip()
        if not raw:
            return default
        try:
            return int(float(raw))  # handles "42.0" from Excel numeric cells
        except ValueError:
            raise ExcelError(
                f"[SETTINGS] key '{key}' must be an integer; got {raw!r}."
            ) from None

    def _bool(key: str, default: bool) -> bool:
        raw = settings.get(key, "").strip().lower()
        if not raw:
            return default
        if raw in ("true", "yes", "1"):
            return True
        if raw in ("false", "no", "0"):
            return False
        raise ExcelError(
            f"[SETTINGS] key '{key}' must be true/false; got {raw!r}."
        )

    alpha        = _float("alpha",        0.05)
    power_target = _float("power",        0.80)
    sigma        = _float("sigma",        1.0)
    r2_target    = _float("r2_target",    0.25)
    max_n        = _int("max_n",          500)
    criterion    = settings.get("criterion", "I").strip() or "I"
    starts       = _int("starts",         5)
    max_iter_do  = _int("max_iter",       1000)
    random_state = _int("random_state",   123)
    # Blocked design options (Enhancement 20)
    n_blocks_raw          = _int("n_blocks",          0)
    block_factor_name_raw = settings.get("block_factor_name", "Block").strip() or "Block"
    # Categorical pre-allocation options (Enhancement 26)
    preallocate_categorical = _bool("preallocate_categorical", False)
    alloc_min_per_cell      = _int("alloc_min_per_cell",  1)
    alloc_max_per_cell_raw  = _int("alloc_max_per_cell",  0)  # 0 → None (no limit)
    # Split-plot options (Enhancement 22)
    htc_factors_raw    = str(settings.get("htc_factors", "") or "").strip()
    n_whole_plots_raw  = _int("n_whole_plots", 0)
    sp_eta             = _float("eta",            1.0)
    subplots_per_wp_raw = _int("subplots_per_wp", 0)   # 0 → auto
    df_method_sp       = str(settings.get("df_method", "auto") or "auto").strip() or "auto"
    # GLM options (GL-9)
    glm_family   = str(settings.get("family", "binomial") or "binomial").strip() or "binomial"
    glm_link     = str(settings.get("link", "") or "").strip() or None
    glm_baseline = _float("baseline", 0.0)

    _do_d: Dict[str, Any] = dict(
        criterion=criterion,
        starts=starts,
        max_iter=max_iter_do,
        random_state=random_state,
        n_blocks=n_blocks_raw,
        block_factor_name=block_factor_name_raw,
        preallocate_categorical=preallocate_categorical,
        alloc_min_per_cell=alloc_min_per_cell,
        alloc_max_per_cell=alloc_max_per_cell_raw,
    )
    if htc_factors_raw and n_whole_plots_raw >= 2:
        htc_list = [f.strip() for f in htc_factors_raw.split(",") if f.strip()]
        if htc_list:
            _do_d["split_plot"] = dict(
                htc_factors=htc_list,
                n_whole_plots=n_whole_plots_raw,
                eta=sp_eta,
                subplots_per_wp=subplots_per_wp_raw,
                df_method=df_method_sp,
            )
    design_opts = build_design_opts(_do_d, error_cls=ExcelError, context="Config sheet")

    # --- Parse [CONTRAST] (if present and single-response contrast/glm mode) ---
    L: Optional[np.ndarray] = None
    delta: Optional[np.ndarray] = None

    if idx_responses is None and power_mode in ("contrast", "glm"):
        if idx_contrast is None:
            raise ExcelError(
                f"power_mode is {power_mode!r} but the '{_SENTINEL_CONTRAST}' sentinel "
                "is missing from the Config sheet."
            )
        contrast_end = idx_factors if idx_factors > idx_contrast else len(rows)
        l_rows: List[List[float]] = []
        delta_values: Optional[List[float]] = None

        for row in rows[idx_contrast + 1: contrast_end]:
            col_a = row[0] if len(row) > 0 else ""
            col_b = row[1] if len(row) > 1 else ""
            if not col_a:
                continue
            if col_a == "L_row":
                try:
                    l_rows.append([float(v.strip()) for v in col_b.split(",") if v.strip()])
                except ValueError as e:
                    raise ExcelError(
                        f"[CONTRAST] L_row contains non-numeric value: {col_b!r}."
                    ) from e
            elif col_a == "delta":
                try:
                    delta_values = [float(v.strip()) for v in col_b.split(",") if v.strip()]
                except ValueError as e:
                    raise ExcelError(
                        f"[CONTRAST] delta contains non-numeric value: {col_b!r}."
                    ) from e

        if not l_rows:
            raise ExcelError("[CONTRAST] has no 'L_row' entries.")
        if delta_values is None:
            raise ExcelError("[CONTRAST] is missing the 'delta' row.")
        if len(delta_values) != len(l_rows):
            raise ExcelError(
                f"[CONTRAST] delta has {len(delta_values)} value(s) but "
                f"there are {len(l_rows)} L_row(s). They must have the same length."
            )
        L = np.array(l_rows, dtype=float)
        delta = np.array(delta_values, dtype=float)

    # --- Parse [FACTORS] ---
    factors: Dict[str, Any] = {}
    factors_data_start = idx_factors + 2  # +1 sentinel, +1 header
    # Stop at [RESPONSES] if present
    factors_end = idx_responses if idx_responses is not None and idx_responses > idx_factors else len(rows)

    for row in rows[factors_data_start:factors_end]:
        col_a = row[0] if len(row) > 0 else ""
        col_b = row[1] if len(row) > 1 else ""
        values = [row[c] for c in range(2, len(row)) if row[c]]

        if not col_a:
            continue

        factor_type = col_b.lower()
        if factor_type == "continuous":
            if len(values) < 2:
                raise ExcelError(
                    f"Factor '{col_a}' is continuous but has fewer than 2 values."
                )
            try:
                factors[col_a] = (float(values[0]), float(values[1]))
            except ValueError as e:
                raise ExcelError(
                    f"Factor '{col_a}' continuous bounds must be numeric; "
                    f"got {values[0]!r}, {values[1]!r}."
                ) from e
        elif factor_type == "categorical":
            if len(values) < 2:
                raise ExcelError(
                    f"Factor '{col_a}' is categorical but has fewer than 2 levels."
                )
            factors[col_a] = values
        else:
            raise ExcelError(
                f"Factor '{col_a}' has unrecognised type {col_b!r}. "
                "Must be 'continuous' or 'categorical'."
            )

    if not factors:
        raise ExcelError("[FACTORS] section contains no factor definitions.")

    # --- Build power_cfg (single-response mode only) ---
    power_cfg: Optional[Union[PowerContrastConfig, PowerR2Config, PowerGLMContrastConfig]] = None
    if idx_responses is None:
        if power_mode == "r2":
            power_cfg = build_power_cfg(
                dict(power_mode="r2", r2_target=r2_target, alpha=alpha,
                     power=power_target, sigma=sigma, max_n=max_n),
                error_cls=ExcelError, context="Config sheet",
            )
        elif power_mode == "glm":
            if not glm_baseline:
                raise ExcelError(
                    "[SETTINGS] 'baseline' is required when power_mode is 'glm'."
                )
            power_cfg = build_power_cfg(
                dict(power_mode="glm", L=L.tolist() if L is not None else None,
                     delta=delta.tolist() if delta is not None else None,
                     baseline=glm_baseline, family=glm_family, link=glm_link,
                     alpha=alpha, power=power_target, max_n=max_n),
                error_cls=ExcelError, context="Config sheet",
            )
        else:
            power_cfg = build_power_cfg(
                dict(power_mode="contrast", L=L.tolist() if L is not None else None,
                     delta=delta.tolist() if delta is not None else None,
                     alpha=alpha, power=power_target, sigma=sigma, max_n=max_n),
                error_cls=ExcelError, context="Config sheet",
            )

    # --- Parse [RESPONSES] section (optional — multi-response mode) ---
    multi_cfg: Optional[MultiResponseOptions] = None

    if idx_responses is not None:
        # Column order: name|power_mode|sigma|alpha|power|weight|L_row|delta|r2_target|
        #               formula|lambda_mode|max_n|max_iter|tol_power
        # Special key rows (interleaved): power_combination, sigma_joint
        responses_data_start = idx_responses + 2
        specs: List[ResponseSpec] = []
        power_combination = "min"
        sigma_joint_arr: Optional[np.ndarray] = None

        for row in rows[responses_data_start:]:
            col_a = str(row[0]) if len(row) > 0 and row[0] is not None else ""
            col_b = str(row[1]) if len(row) > 1 and row[1] is not None else ""
            if not col_a:
                continue
            if col_a == "power_combination":
                power_combination = col_b or "min"
                continue
            if col_a == "sigma_joint":
                # Format: semicolon-separated matrix rows, comma-separated values.
                if col_b:
                    try:
                        _sj_rows = []
                        for _line in col_b.split(";"):
                            _line = _line.strip()
                            if _line:
                                _sj_rows.append([float(x) for x in _line.replace(",", " ").split()])
                        sigma_joint_arr = np.array(_sj_rows, dtype=float)
                    except (ValueError, TypeError) as e:
                        raise ExcelError(
                            f"[RESPONSES] sigma_joint: invalid matrix format: {e}"
                        ) from e
                continue

            def _rcell(idx: int) -> str:
                v = row[idx] if len(row) > idx else None
                return str(v) if v is not None else ""

            r_name = col_a
            r_mode = col_b.lower()
            r_sigma = float(_rcell(2)) if _rcell(2) else sigma
            r_alpha = float(_rcell(3)) if _rcell(3) else alpha
            r_power_v = float(_rcell(4)) if _rcell(4) else power_target
            r_weight = float(_rcell(5)) if _rcell(5) else 1.0
            r_L_str  = _rcell(6)
            r_d_str  = _rcell(7)
            r_r2_str = _rcell(8)
            r_formula_str = _rcell(9) or None
            r_lambda_mode = _rcell(10) or "n"
            r_max_n   = int(float(_rcell(11))) if _rcell(11) else 2000
            r_max_iter = int(float(_rcell(12))) if _rcell(12) else 200
            r_tol_power = float(_rcell(13)) if _rcell(13) else 1e-3
            # cols 14/15: GLM family and baseline (optional — blank for non-GLM responses)
            r_glm_family = _rcell(14).strip() or None
            r_glm_baseline_str = _rcell(15).strip()

            _resp_ctx = f"[RESPONSES] row '{r_name}'"
            if r_mode == "contrast":
                if not r_L_str or not r_d_str:
                    raise ExcelError(
                        f"{_resp_ctx}: contrast mode requires "
                        "L_row (col 7) and delta (col 8) values."
                    )
                try:
                    L_vals = [float(v.strip()) for v in r_L_str.split(",") if v.strip()]
                    d_vals = [float(v.strip()) for v in r_d_str.split(",") if v.strip()]
                except ValueError as e:
                    raise ExcelError(
                        f"{_resp_ctx}: non-numeric L_row or delta: {e}"
                    ) from e
                pcfg: Union[PowerContrastConfig, PowerR2Config, PowerGLMContrastConfig] = (
                    build_power_cfg(
                        dict(power_mode="contrast", L=[L_vals], delta=d_vals,
                             alpha=r_alpha, power=r_power_v, sigma=r_sigma,
                             max_n=r_max_n, max_iter=r_max_iter, tol_power=r_tol_power),
                        error_cls=ExcelError, context=_resp_ctx,
                    )
                )
            elif r_mode == "r2":
                if not r_r2_str:
                    raise ExcelError(
                        f"{_resp_ctx}: r2 mode requires r2_target (col 9)."
                    )
                pcfg = build_power_cfg(
                    dict(power_mode="r2", r2_target=float(r_r2_str),
                         alpha=r_alpha, power=r_power_v, sigma=r_sigma,
                         lambda_mode=r_lambda_mode,
                         max_n=r_max_n, max_iter=r_max_iter, tol_power=r_tol_power),
                    error_cls=ExcelError, context=_resp_ctx,
                )
            elif r_mode == "glm":
                if not r_L_str or not r_d_str:
                    raise ExcelError(
                        f"{_resp_ctx}: glm mode requires "
                        "L_row (col 7) and delta (col 8) values."
                    )
                if not r_glm_baseline_str:
                    raise ExcelError(
                        f"{_resp_ctx}: glm mode requires baseline (col 16)."
                    )
                try:
                    L_vals = [float(v.strip()) for v in r_L_str.split(",") if v.strip()]
                    d_vals = [float(v.strip()) for v in r_d_str.split(",") if v.strip()]
                    r_glm_baseline = float(r_glm_baseline_str)
                except ValueError as e:
                    raise ExcelError(
                        f"{_resp_ctx}: non-numeric value: {e}"
                    ) from e
                pcfg = build_power_cfg(
                    dict(power_mode="glm", L=[L_vals], delta=d_vals,
                         baseline=r_glm_baseline, family=r_glm_family or "binomial",
                         alpha=r_alpha, power=r_power_v, max_n=r_max_n),
                    error_cls=ExcelError, context=_resp_ctx,
                )
            else:
                raise ExcelError(
                    f"{_resp_ctx}: power_mode must be 'contrast', 'r2', or 'glm'; "
                    f"got {col_b!r}."
                )

            try:
                specs.append(ResponseSpec(
                    name=r_name, power_cfg=pcfg,
                    formula=r_formula_str, weight=r_weight,
                ))
            except (ValueError, TypeError) as e:
                raise ExcelError(f"{_resp_ctx}: {e}") from e

        if len(specs) < 2:
            raise ExcelError(
                f"[RESPONSES] section must define at least 2 responses; found {len(specs)}."
            )
        try:
            multi_cfg = MultiResponseOptions(
                responses=specs,
                power_combination=power_combination,
                sigma_joint=sigma_joint_arr,
            )
        except (ValueError, TypeError) as e:
            raise ExcelError(f"Config sheet produced invalid MultiResponseOptions: {e}") from e

    return formula, factors, power_cfg, design_opts, multi_cfg


# ---------------------------------------------------------------------------
# Output sheet writers
# ---------------------------------------------------------------------------
def _write_df_to_sheet(wb: Any, sheet_name: str, df: pd.DataFrame) -> None:
    """Write *df* to a sheet in *wb*, creating or clearing it as needed."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    _header_row(ws, 1, list(df.columns))
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = ""
            ws.cell(row=r_idx, column=c_idx, value=val)
    _set_column_widths(ws, [max(len(str(c)), 10) + 2 for c in df.columns])



def _write_output_sheets(
    wb: Any,
    result: Dict[str, Any],
    report: Dict[str, Any],
    design_df: pd.DataFrame,
    buckets_df: pd.DataFrame,
) -> None:
    """Write every result sheet into *wb* (shared by run and tests).

    Beyond Results/Design/Buckets, this carries the coding authorities:
    ``ModelMatrix`` is the basis the power calculation used (UX-57 — for
    data-dependent codings a refit of the Design sheet re-learns spline
    knots), and for COMPOUND multi-response runs one sheet per response
    (UX-63/UX-66), because each response was powered on its own formula's
    basis and a data-dependent response may not be reproducible from any
    other sheet. Sheet titles cap at 31 characters and forbid several
    separators, so free-form response names are slugged; the
    ``ModelMatrixIndex`` sheet maps original names to sheet titles.

    Output sheets always describe THIS result only: per-response sheets a
    previous export recorded in ``ModelMatrixIndex`` are deleted up front
    (UX-73). Without that, a repeat export whose response names changed —
    even by case only, which Excel titles do not distinguish — leaves the
    old sheet in place, the backend silently renames the new one around it,
    and the rewritten index points at a sheet that does not exist.
    """
    _reconcile_previous_matrix_sheets(wb)
    _write_results_sheet(wb, report)
    _write_df_to_sheet(wb, "Design", design_df)
    _write_df_to_sheet(wb, "Buckets", buckets_df)
    if result.get("model_matrix") is not None:
        _write_df_to_sheet(wb, "ModelMatrix", result["model_matrix"])
    elif "ModelMatrix" in wb.sheetnames:
        del wb["ModelMatrix"]  # do not pair the new Design with an old basis
    if (
        result.get("model_matrices") is not None
        and report.get("compound_criterion")
    ):
        # Collisions (with the user's own sheets — ours are already gone)
        # are checked on the COMPLETE title: openpyxl compares titles
        # case-insensitively, so a bare-slug check lets it rename the sheet
        # while the index records the requested title (UX-73).
        _mm_taken: set = set(wb.sheetnames) | {"ModelMatrixIndex"}
        _mm_index_rows = []
        for _rname, _rmm in result["model_matrices"].items():
            _sheet = "MM_" + safe_name_slug(
                _rname, _mm_taken, maxlen=27, prefix="MM_"
            )
            _write_df_to_sheet(wb, _sheet, _rmm)
            _mm_index_rows.append((_rname, _sheet))
        _write_df_to_sheet(
            wb, "ModelMatrixIndex",
            pd.DataFrame(_mm_index_rows, columns=["response", "sheet"]),
        )


def _reconcile_previous_matrix_sheets(wb: Any) -> None:
    """Delete per-response sheets a previous export listed in its index.

    Only titles the index attributes to us (``MM_`` prefix) are deleted —
    the index is the record of what WE wrote, so the user's own sheets
    survive even if an edited index names them. The index sheet itself is
    removed too; a compound result rewrites it, any other result must not
    leave it dangling.
    """
    if "ModelMatrixIndex" not in wb.sheetnames:
        return
    ws_idx = wb["ModelMatrixIndex"]
    for _row in ws_idx.iter_rows(min_row=2, max_col=2, values_only=True):
        _title = _row[1] if len(_row) > 1 else None
        if (
            isinstance(_title, str)
            and _title.startswith("MM_")
            and _title in wb.sheetnames
        ):
            del wb[_title]
    del wb["ModelMatrixIndex"]


def _write_results_sheet(wb: Any, report: Dict[str, Any]) -> None:
    """Write the run report dict as a key/value table to the Results sheet."""
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")
    _header_row(ws, 1, ["Key", "Value"])
    r = 2
    for key, val in report.items():
        if isinstance(val, dict):  # flatten nested dicts (e.g. diagnostics)
            for sub_key, sub_val in val.items():
                if isinstance(sub_val, list):
                    sub_val = str(sub_val)
                ws.cell(row=r, column=1, value=f"{key}.{sub_key}").font = Font(bold=True)
                ws.cell(row=r, column=2, value=sub_val)
                r += 1
        elif isinstance(val, list):
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=str(val))
            r += 1
        else:
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Template builder
# ---------------------------------------------------------------------------
def create_excel_template(
    path: Union[str, Path] = "iopt_template.xlsx",
    example: str = "r2",
) -> Path:
    """Create a new ``.xlsx`` workbook pre-populated with a runnable example.

    The workbook contains a **Config** sheet that ``excel_run()`` can read
    directly, plus empty placeholder sheets for Results, Design, and Buckets
    that ``excel_run()`` will populate after the design search.

    Dropdown validation is applied to the ``power_mode`` and ``criterion``
    cells.

    Parameters
    ----------
    path : str or Path, default ``"iopt_template.xlsx"``
        Destination file path.  Existing files are overwritten.
    example : {"r2", "contrast", "multiresponse", "glm-binomial", "glm-poisson"}, default ``"r2"``
        Which pre-filled example to write.

        * ``"r2"``            — global R² power config with two continuous factors.
        * ``"contrast"``      — single contrast (L matrix + delta) with two
          continuous factors.
        * ``"multiresponse"`` — two-response R² joint design with a ``[RESPONSES]``
          section demonstrating all per-response and ``sigma_joint`` fields.
        * ``"glm-binomial"``  — GLM logistic power with Wald χ² test (binomial family).
        * ``"glm-poisson"``   — GLM log-linear power with Wald χ² test (Poisson family).

    Returns
    -------
    Path
        Absolute path of the created workbook.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    ValueError
        If *example* is not one of the supported values.
    """
    _require_openpyxl()

    if example not in ("r2", "contrast", "multiresponse", "glm-binomial", "glm-poisson"):
        raise ValueError(
            f"Unknown example {example!r}. Supported values: "
            "'r2', 'contrast', 'multiresponse', 'glm-binomial', 'glm-poisson'."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Config"

    r = 1  # current row counter

    # ------------------------------------------------------------------ #
    # [SETTINGS]
    # ------------------------------------------------------------------ #
    _sentinel_style(ws, r, _SENTINEL_SETTINGS); r += 1

    _key_cell(ws, r, "formula", "~ 1 + A + B"); r += 1
    # power_mode is omitted for multiresponse (overridden by [RESPONSES] section)
    if example != "multiresponse":
        # GLM examples use power_mode="glm"; map template name to mode value
        _pm_value = "glm" if example.startswith("glm") else example
        _key_cell(ws, r, "power_mode", _pm_value); r += 1
        _add_dropdown(ws, '"r2,contrast,glm"', row=r - 1)
        _key_cell(ws, r, "alpha",        0.05);  r += 1
        _key_cell(ws, r, "power",        0.80);  r += 1
        if example == "r2":
            _key_cell(ws, r, "r2_target",  0.25); r += 1
            _key_cell(ws, r, "sigma",      "");   r += 1   # not used in R² mode
        elif example.startswith("glm"):
            _glm_family   = "binomial" if example == "glm-binomial" else "poisson"
            _glm_baseline = 0.20       if example == "glm-binomial" else 2.0
            _key_cell(ws, r, "family",    _glm_family);   r += 1
            _key_cell(ws, r, "link",      "");             r += 1  # blank = canonical
            _key_cell(ws, r, "baseline",  _glm_baseline); r += 1
        else:
            _key_cell(ws, r, "sigma",      1.0);  r += 1
            _key_cell(ws, r, "r2_target",  "");   r += 1   # not used in contrast mode
        _key_cell(ws, r, "max_n",        500);   r += 1

    _key_cell(ws, r, "criterion",    "I");   r += 1
    _add_dropdown(ws, '"I,D,A"', row=r - 1)

    _key_cell(ws, r, "starts",       5);     r += 1
    _key_cell(ws, r, "max_iter",     1000);  r += 1
    _key_cell(ws, r, "random_state", 123);   r += 1
    # Blocked design: set n_blocks >= 2 to enable; 0 = unblocked (default).
    _key_cell(ws, r, "n_blocks",                0);       r += 1
    _key_cell(ws, r, "block_factor_name",       "Block"); r += 1
    # Categorical pre-allocation: set to true/false.
    _key_cell(ws, r, "preallocate_categorical", "false"); r += 1
    _add_dropdown(ws, '"true,false"', row=r - 1)
    _key_cell(ws, r, "alloc_min_per_cell",      1);       r += 1
    _key_cell(ws, r, "alloc_max_per_cell",      0);       r += 1  # 0 = no upper limit
    # Split-plot: set htc_factors and n_whole_plots >= 2 to enable.
    _key_cell(ws, r, "htc_factors",             "");      r += 1  # comma-separated HTC factor names
    _key_cell(ws, r, "n_whole_plots",           0);       r += 1  # 0 = disabled; >= 2 to enable
    _key_cell(ws, r, "eta",                     1.0);     r += 1  # variance ratio sigma2_wp/sigma2_sp
    _key_cell(ws, r, "subplots_per_wp",         0);       r += 1  # 0 = auto
    _key_cell(ws, r, "df_method",               "auto");  r += 1
    _add_dropdown(ws, '"auto,conservative,sp_only"', row=r - 1)

    r += 1  # blank separator

    # ------------------------------------------------------------------ #
    # [CONTRAST] — only for contrast example
    # ------------------------------------------------------------------ #
    if example == "contrast":
        _sentinel_style(ws, r, _SENTINEL_CONTRAST); r += 1
        # Contrast between A=-1 and A=+1, holding B at mid-range
        # Model: Intercept + A + B  → L = [0, 1, 0]
        _key_cell(ws, r, "L_row",  "0, 1, 0"); r += 1
        _key_cell(ws, r, "delta",  "1.0");      r += 1
        r += 1  # blank separator
    elif example == "glm-binomial":
        _sentinel_style(ws, r, _SENTINEL_CONTRAST); r += 1
        _key_cell(ws, r, "L_row",  "0, 1, 0"); r += 1
        _key_cell(ws, r, "delta",  "0.5");      r += 1
        r += 1
    elif example == "glm-poisson":
        _sentinel_style(ws, r, _SENTINEL_CONTRAST); r += 1
        _key_cell(ws, r, "L_row",  "0, 1, 0"); r += 1
        _key_cell(ws, r, "delta",  "0.3");      r += 1
        r += 1

    # ------------------------------------------------------------------ #
    # [FACTORS]
    # ------------------------------------------------------------------ #
    _sentinel_style(ws, r, _SENTINEL_FACTORS); r += 1
    _header_row(ws, r, ["Name", "Type", "Value 1", "Value 2"]); r += 1
    ws.cell(row=r, column=1, value="A")
    ws.cell(row=r, column=2, value="continuous")
    ws.cell(row=r, column=3, value=-1.0)
    ws.cell(row=r, column=4, value=1.0)
    r += 1
    ws.cell(row=r, column=1, value="B")
    ws.cell(row=r, column=2, value="continuous")
    ws.cell(row=r, column=3, value=-1.0)
    ws.cell(row=r, column=4, value=1.0)
    r += 1

    # ------------------------------------------------------------------ #
    # [RESPONSES] — only for multiresponse example
    # Column order: name|power_mode|sigma|alpha|power|weight|L_row|delta|
    #               r2_target|formula|lambda_mode|max_n|max_iter|tol_power
    # ------------------------------------------------------------------ #
    if example == "multiresponse":
        r += 1  # blank separator
        _sentinel_style(ws, r, _SENTINEL_RESPONSES); r += 1
        # Header row (informational — not parsed by _read_config_sheet)
        _header_row(ws, r, [
            "name", "power_mode", "sigma", "alpha", "power", "weight",
            "L_row", "delta", "r2_target", "formula",
            "lambda_mode", "max_n", "max_iter", "tol_power",
        ]); r += 1
        # Special key rows
        _key_cell(ws, r, "power_combination", "min"); r += 1
        # sigma_joint: leave blank to disable; fill as "1,0.3; 0.3,1" to enable Hotelling T²
        _key_cell(ws, r, "sigma_joint", ""); r += 1
        # Per-response data rows
        for _rname, _r2t in (("Y1", 0.15), ("Y2", 0.20)):
            ws.cell(row=r, column=1, value=_rname)
            ws.cell(row=r, column=2, value="r2")
            # sigma/alpha/power/weight — blank = use global defaults
            ws.cell(row=r, column=6, value=1.0)   # weight
            ws.cell(row=r, column=9, value=_r2t)  # r2_target
            ws.cell(row=r, column=11, value="n")   # lambda_mode
            r += 1

    _set_column_widths(ws, [20, 16, 12, 12])

    # ------------------------------------------------------------------ #
    # Placeholder output sheets
    # ------------------------------------------------------------------ #
    for sheet in ("Results", "Design", "Buckets"):
        out_ws = wb.create_sheet(sheet)
        out_ws.cell(row=1, column=1, value=f"Run excel_run() to populate this sheet.")
        out_ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    dest = Path(path).resolve()
    wb.save(dest)
    return dest


# ---------------------------------------------------------------------------
# Run function
# ---------------------------------------------------------------------------
def excel_run(
    path: Union[str, Path],
    design_opts_override: Optional[DesignOptions] = None,
) -> Dict[str, Any]:
    """Read a Config sheet from an ``.xlsx`` workbook, run the design search,
    and write Results / Design / Buckets sheets back into the **same** file.

    Parameters
    ----------
    path : str or Path
        Path to the ``.xlsx`` workbook (must contain a ``Config`` sheet).
    design_opts_override : DesignOptions, optional
        If provided, these options override the values read from the Config
        sheet.  Useful for programmatic control without editing the file.

    Returns
    -------
    dict
        The full result dict returned by ``find_optimal_design``, with
        an additional ``"excel_path"`` key containing the absolute workbook
        path.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    ExcelError
        If the workbook cannot be opened, the Config sheet is missing or
        malformed, or the design run fails.
    """
    _require_openpyxl()
    from .api import find_optimal_design

    path = Path(path).resolve()
    if not path.exists():
        raise ExcelError(f"Workbook not found: {path}")

    # ------------------------------------------------------------------
    # 1. Open and parse Config sheet
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        raise ExcelError(f"Could not open workbook {path}: {e}") from e

    if "Config" not in wb.sheetnames:
        raise ExcelError(
            f"Workbook {path.name!r} has no 'Config' sheet. "
            "Use create_excel_template() to create a properly structured workbook."
        )

    try:
        formula, factors, power_cfg, design_opts, multi_cfg = _read_config_sheet(wb["Config"])
    except ExcelError:
        raise
    except Exception as e:
        raise ExcelError(f"Failed to parse Config sheet: {e}") from e

    if design_opts_override is not None:
        design_opts = design_opts_override

    # ------------------------------------------------------------------
    # 2. Run design search
    # ------------------------------------------------------------------
    try:
        if multi_cfg is not None:
            from .api import find_multiresponse_design  # noqa: PLC0415
            result = find_multiresponse_design(
                formula=formula,
                factors=factors,
                multi_cfg=multi_cfg,
                design_opts=design_opts,
            )
        else:
            result = find_optimal_design(
                formula=formula,
                factors=factors,
                power_cfg=power_cfg,
                design_opts=design_opts,
            )
    except Exception as e:
        raise ExcelError(f"Design search failed: {e}") from e

    # Unified envelope for both modes (UX-6).
    design_df: pd.DataFrame = result["design_df"]
    buckets_df: pd.DataFrame = result["buckets_df"]
    report: Dict[str, Any] = result["report"]
    if multi_cfg is not None:
        # Flatten per-response powers for the results sheet.
        for _r in report.get("responses", []):
            report[f"{_r['name']}_power"] = _r["power"]

    # ------------------------------------------------------------------
    # 3. Write output sheets back into the workbook
    # ------------------------------------------------------------------
    try:
        _write_output_sheets(wb, result, report, design_df, buckets_df)
        wb.save(path)
    except Exception as e:
        raise ExcelError(f"Failed to write results to workbook: {e}") from e

    result["excel_path"] = str(path)
    return result


# ---------------------------------------------------------------------------
# Public exports
# ---------------------------------------------------------------------------
__all__ = ["ExcelError", "create_excel_template", "excel_run"]
