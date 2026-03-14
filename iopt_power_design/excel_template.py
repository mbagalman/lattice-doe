# excel_template.py
# License: MIT
"""
Excel workbook bidirectional connector for iopt-power-design
=============================================================

Creates a structured ``.xlsx`` workbook pre-filled with a runnable example
and, after a design run, writes Results / Design / Buckets sheets back into
the **same** file.

Sheet layout (Config sheet)
---------------------------
The ``Config`` sheet uses the same sentinel-header structure as the Google
Sheets connector (``sheets.py``), so the same mental model applies:

  ``[SETTINGS]``  — key/value pairs (formula, power parameters, design options).
  ``[CONTRAST]``  — optional; required when ``power_mode = contrast``.
  ``[FACTORS]``   — factor table: Name | Type | Value 1 | Value 2 | …

Dropdown validation is applied to ``power_mode`` (r2 / contrast) and
``criterion`` (I / D / A) cells.

Output sheets
-------------
``excel_run()`` writes three sheets into the workbook after a successful run:

  ``Results``  — summary key/value table (n, power, elapsed, …)
  ``Design``   — full design DataFrame
  ``Buckets``  — run-frequency buckets

Install
-------
  pip install "iopt-power-design[extras]"
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

from .config import (
    PowerContrastConfig, PowerR2Config, DesignOptions, SplitPlotOptions,
    MultiResponseOptions, ResponseSpec,
)

# ---------------------------------------------------------------------------
# Soft dependency guard
# ---------------------------------------------------------------------------
try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_INSTALL_HINT = 'pip install "iopt-power-design[extras]"'

# ---------------------------------------------------------------------------
# Section sentinels (must match sheets.py constants for consistency)
# ---------------------------------------------------------------------------
_SENTINEL_SETTINGS  = "[SETTINGS]"
_SENTINEL_CONTRAST  = "[CONTRAST]"
_SENTINEL_FACTORS   = "[FACTORS]"
_SENTINEL_RESPONSES = "[RESPONSES]"

# ---------------------------------------------------------------------------
# Custom exception
# ---------------------------------------------------------------------------
class ExcelError(RuntimeError):
    """Raised for all Excel integration failures.

    Covers missing sections, malformed data, parse errors, and write failures.
    The underlying exception is attached as ``__cause__``.
    """


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise ImportError(
            f"openpyxl is required for Excel integration. {_INSTALL_HINT}"
        )


def _sentinel_style(ws: Any, row: int, label: str) -> None:
    """Apply bold dark-blue styling to a sentinel header row."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="left")


def _key_cell(ws: Any, row: int, key: str, value: Any = "") -> None:
    """Write a key in column A (bold) and value in column B."""
    kc = ws.cell(row=row, column=1, value=key)
    kc.font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def _header_row(ws: Any, row: int, headers: List[str]) -> None:
    """Write column headers in bold with a light-blue background."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")


def _set_column_widths(ws: Any, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_dropdown(ws: Any, formula: str, row: int, col: int = 2) -> None:
    """Attach an in-cell dropdown DataValidation to the given cell."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=False,
                        showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=col))


# ---------------------------------------------------------------------------
# Config sheet reader (openpyxl-based, mirrors sheets.py._parse_config_sheet)
# ---------------------------------------------------------------------------
def _read_config_sheet(
    ws: Any,
) -> Tuple[str, Dict[str, Any], Optional[Union[PowerContrastConfig, PowerR2Config]], DesignOptions, Optional[MultiResponseOptions]]:
    """Parse the Config worksheet and return ``(formula, factors, power_cfg, design_opts, multi_cfg)``.

    Raises
    ------
    ExcelError
        For missing sentinels, missing required keys, or malformed data.
    """
    # Collect all rows as lists of stripped strings
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])

    # Locate sentinels
    idx_settings:  Optional[int] = None
    idx_contrast:  Optional[int] = None
    idx_factors:   Optional[int] = None
    idx_responses: Optional[int] = None

    for i, row in enumerate(rows):
        col_a = row[0] if row else ""
        if col_a == _SENTINEL_SETTINGS:
            idx_settings = i
        elif col_a == _SENTINEL_CONTRAST:
            idx_contrast = i
        elif col_a == _SENTINEL_FACTORS:
            idx_factors = i
        elif col_a == _SENTINEL_RESPONSES:
            idx_responses = i

    if idx_settings is None:
        raise ExcelError(
            f"Config sheet is missing the '{_SENTINEL_SETTINGS}' sentinel. "
            "Ensure cell A1 (or equivalent) contains '[SETTINGS]'."
        )
    if idx_factors is None:
        raise ExcelError(
            f"Config sheet is missing the '{_SENTINEL_FACTORS}' sentinel."
        )

    # --- Parse [SETTINGS] ---
    sentinels_after = sorted(
        j for j in (idx_contrast, idx_factors, idx_responses)
        if j is not None and j > idx_settings
    )
    settings_end = sentinels_after[0] if sentinels_after else len(rows)

    settings: Dict[str, str] = {}
    for row in rows[idx_settings + 1: settings_end]:
        col_a = row[0] if len(row) > 0 else ""
        col_b = row[1] if len(row) > 1 else ""
        if col_a:
            settings[col_a] = col_b

    if not settings.get("formula"):
        raise ExcelError(
            "Config sheet [SETTINGS] is missing the required 'formula' key."
        )
    # power_mode is required unless a [RESPONSES] section is present
    if idx_responses is None and not settings.get("power_mode"):
        raise ExcelError(
            "Config sheet [SETTINGS] is missing the required 'power_mode' key."
        )

    formula = settings["formula"]
    power_mode = settings.get("power_mode", "").strip().lower()

    if power_mode and power_mode not in ("r2", "contrast"):
        raise ExcelError(
            f"[SETTINGS] power_mode must be 'r2' or 'contrast'; got {settings.get('power_mode')!r}."
        )

    def _float(key: str, default: float) -> float:
        raw = settings.get(key, "").strip()
        if not raw:
            return default
        try:
            return float(raw)
        except ValueError:
            raise ExcelError(
                f"[SETTINGS] key '{key}' must be a number; got {raw!r}."
            ) from None

    def _int(key: str, default: int) -> int:
        raw = settings.get(key, "").strip()
        if not raw:
            return default
        try:
            return int(float(raw))  # handles "42.0" from Excel numeric cells
        except ValueError:
            raise ExcelError(
                f"[SETTINGS] key '{key}' must be an integer; got {raw!r}."
            ) from None

    def _bool(key: str, default: bool) -> bool:
        raw = settings.get(key, "").strip().lower()
        if not raw:
            return default
        if raw in ("true", "yes", "1"):
            return True
        if raw in ("false", "no", "0"):
            return False
        raise ExcelError(
            f"[SETTINGS] key '{key}' must be true/false; got {raw!r}."
        )

    alpha        = _float("alpha",        0.05)
    power_target = _float("power",        0.80)
    sigma        = _float("sigma",        1.0)
    r2_target    = _float("r2_target",    0.25)
    max_n        = _int("max_n",          500)
    criterion    = settings.get("criterion", "I").strip() or "I"
    starts       = _int("starts",         5)
    max_iter_do  = _int("max_iter",       1000)
    random_state = _int("random_state",   123)
    # Blocked design options (Enhancement 20)
    n_blocks_raw          = _int("n_blocks",          0)
    block_factor_name_raw = settings.get("block_factor_name", "Block").strip() or "Block"
    # Categorical pre-allocation options (Enhancement 26)
    preallocate_categorical = _bool("preallocate_categorical", False)
    alloc_min_per_cell      = _int("alloc_min_per_cell",  1)
    alloc_max_per_cell_raw  = _int("alloc_max_per_cell",  0)  # 0 → None (no limit)
    # Split-plot options (Enhancement 22)
    htc_factors_raw    = str(settings.get("htc_factors", "") or "").strip()
    n_whole_plots_raw  = _int("n_whole_plots", 0)
    sp_eta             = _float("eta",            1.0)
    subplots_per_wp_raw = _int("subplots_per_wp", 0)   # 0 → auto
    df_method_sp       = str(settings.get("df_method", "auto") or "auto").strip() or "auto"

    do_kwargs: Dict[str, Any] = dict(
        criterion=criterion,
        starts=starts,
        max_iter=max_iter_do,
        random_state=random_state,
    )
    if n_blocks_raw >= 2:
        do_kwargs["n_blocks"] = n_blocks_raw
        do_kwargs["block_factor_name"] = block_factor_name_raw
    if preallocate_categorical:
        do_kwargs["preallocate_categorical"] = True
        do_kwargs["alloc_min_per_cell"] = alloc_min_per_cell
        if alloc_max_per_cell_raw > 0:
            do_kwargs["alloc_max_per_cell"] = alloc_max_per_cell_raw
    # Split-plot (Enhancement 22)
    if htc_factors_raw and n_whole_plots_raw >= 2:
        htc_list = [f.strip() for f in htc_factors_raw.split(",") if f.strip()]
        if htc_list:
            try:
                do_kwargs["split_plot"] = SplitPlotOptions(
                    htc_factors=htc_list,
                    n_whole_plots=n_whole_plots_raw,
                    eta=sp_eta,
                    subplots_per_wp=subplots_per_wp_raw if subplots_per_wp_raw > 0 else None,
                    df_method=df_method_sp,
                )
            except (ValueError, TypeError) as e:
                raise ExcelError(f"Config sheet produced invalid SplitPlotOptions: {e}") from e
    try:
        design_opts = DesignOptions(**do_kwargs)
    except (ValueError, TypeError) as e:
        raise ExcelError(f"Config sheet produced invalid DesignOptions: {e}") from e

    # --- Parse [CONTRAST] (if present and single-response mode) ---
    L: Optional[np.ndarray] = None
    delta: Optional[np.ndarray] = None

    if idx_responses is None and power_mode == "contrast":
        if idx_contrast is None:
            raise ExcelError(
                f"power_mode is 'contrast' but the '{_SENTINEL_CONTRAST}' sentinel "
                "is missing from the Config sheet."
            )
        contrast_end = idx_factors if idx_factors > idx_contrast else len(rows)
        l_rows: List[List[float]] = []
        delta_values: Optional[List[float]] = None

        for row in rows[idx_contrast + 1: contrast_end]:
            col_a = row[0] if len(row) > 0 else ""
            col_b = row[1] if len(row) > 1 else ""
            if not col_a:
                continue
            if col_a == "L_row":
                try:
                    l_rows.append([float(v.strip()) for v in col_b.split(",") if v.strip()])
                except ValueError as e:
                    raise ExcelError(
                        f"[CONTRAST] L_row contains non-numeric value: {col_b!r}."
                    ) from e
            elif col_a == "delta":
                try:
                    delta_values = [float(v.strip()) for v in col_b.split(",") if v.strip()]
                except ValueError as e:
                    raise ExcelError(
                        f"[CONTRAST] delta contains non-numeric value: {col_b!r}."
                    ) from e

        if not l_rows:
            raise ExcelError("[CONTRAST] has no 'L_row' entries.")
        if delta_values is None:
            raise ExcelError("[CONTRAST] is missing the 'delta' row.")
        if len(delta_values) != len(l_rows):
            raise ExcelError(
                f"[CONTRAST] delta has {len(delta_values)} value(s) but "
                f"there are {len(l_rows)} L_row(s). They must have the same length."
            )
        L = np.array(l_rows, dtype=float)
        delta = np.array(delta_values, dtype=float)

    # --- Parse [FACTORS] ---
    factors: Dict[str, Any] = {}
    factors_data_start = idx_factors + 2  # +1 sentinel, +1 header
    # Stop at [RESPONSES] if present
    factors_end = idx_responses if idx_responses is not None and idx_responses > idx_factors else len(rows)

    for row in rows[factors_data_start:factors_end]:
        col_a = row[0] if len(row) > 0 else ""
        col_b = row[1] if len(row) > 1 else ""
        values = [row[c] for c in range(2, len(row)) if row[c]]

        if not col_a:
            continue

        factor_type = col_b.lower()
        if factor_type == "continuous":
            if len(values) < 2:
                raise ExcelError(
                    f"Factor '{col_a}' is continuous but has fewer than 2 values."
                )
            try:
                factors[col_a] = (float(values[0]), float(values[1]))
            except ValueError as e:
                raise ExcelError(
                    f"Factor '{col_a}' continuous bounds must be numeric; "
                    f"got {values[0]!r}, {values[1]!r}."
                ) from e
        elif factor_type == "categorical":
            if len(values) < 2:
                raise ExcelError(
                    f"Factor '{col_a}' is categorical but has fewer than 2 levels."
                )
            factors[col_a] = values
        else:
            raise ExcelError(
                f"Factor '{col_a}' has unrecognised type {col_b!r}. "
                "Must be 'continuous' or 'categorical'."
            )

    if not factors:
        raise ExcelError("[FACTORS] section contains no factor definitions.")

    # --- Build power_cfg (single-response mode only) ---
    power_cfg: Optional[Union[PowerContrastConfig, PowerR2Config]] = None
    if idx_responses is None:
        if power_mode == "r2":
            try:
                power_cfg = PowerR2Config(
                    r2_target=r2_target,
                    alpha=alpha,
                    power=power_target,
                    sigma=sigma,
                    max_n=max_n,
                )
            except (ValueError, TypeError) as e:
                raise ExcelError(f"Config sheet produced invalid PowerR2Config: {e}") from e
        else:
            try:
                power_cfg = PowerContrastConfig(
                    L=L,
                    delta=delta,
                    alpha=alpha,
                    power=power_target,
                    sigma=sigma,
                    max_n=max_n,
                )
            except (ValueError, TypeError) as e:
                raise ExcelError(f"Config sheet produced invalid PowerContrastConfig: {e}") from e

    # --- Parse [RESPONSES] section (optional — multi-response mode) ---
    multi_cfg: Optional[MultiResponseOptions] = None

    if idx_responses is not None:
        responses_data_start = idx_responses + 2
        specs: List[ResponseSpec] = []
        power_combination = "min"

        for row in rows[responses_data_start:]:
            col_a = row[0] if len(row) > 0 else ""
            col_b = row[1] if len(row) > 1 else ""
            if not col_a:
                continue
            if col_a == "power_combination":
                power_combination = col_b or "min"
                continue

            def _rcell(idx: int) -> str:
                return row[idx] if len(row) > idx else ""

            r_name = col_a
            r_mode = col_b.lower()
            r_sigma = float(_rcell(2)) if _rcell(2) else sigma
            r_alpha = float(_rcell(3)) if _rcell(3) else alpha
            r_power_v = float(_rcell(4)) if _rcell(4) else power_target
            r_weight = float(_rcell(5)) if _rcell(5) else 1.0
            r_L_str  = _rcell(6)
            r_d_str  = _rcell(7)
            r_r2_str = _rcell(8)
            r_formula_str = _rcell(9) or None

            if r_mode == "contrast":
                if not r_L_str or not r_d_str:
                    raise ExcelError(
                        f"[RESPONSES] row '{r_name}': contrast mode requires "
                        "L_row (col 7) and delta (col 8) values."
                    )
                try:
                    L_vals = [float(v.strip()) for v in r_L_str.split(",") if v.strip()]
                    d_vals = [float(v.strip()) for v in r_d_str.split(",") if v.strip()]
                except ValueError as e:
                    raise ExcelError(
                        f"[RESPONSES] row '{r_name}': non-numeric L_row or delta: {e}"
                    ) from e
                r_L = np.array([L_vals], dtype=float)
                r_d = np.array(d_vals, dtype=float)
                try:
                    pcfg: Union[PowerContrastConfig, PowerR2Config] = PowerContrastConfig(
                        L=r_L, delta=r_d, alpha=r_alpha, power=r_power_v, sigma=r_sigma,
                    )
                except (ValueError, TypeError) as e:
                    raise ExcelError(
                        f"[RESPONSES] row '{r_name}': invalid PowerContrastConfig: {e}"
                    ) from e
            elif r_mode == "r2":
                if not r_r2_str:
                    raise ExcelError(
                        f"[RESPONSES] row '{r_name}': r2 mode requires r2_target (col 9)."
                    )
                try:
                    pcfg = PowerR2Config(
                        r2_target=float(r_r2_str),
                        alpha=r_alpha, power=r_power_v, sigma=r_sigma,
                    )
                except (ValueError, TypeError) as e:
                    raise ExcelError(
                        f"[RESPONSES] row '{r_name}': invalid PowerR2Config: {e}"
                    ) from e
            else:
                raise ExcelError(
                    f"[RESPONSES] row '{r_name}': power_mode must be 'contrast' or 'r2'; "
                    f"got {col_b!r}."
                )

            try:
                specs.append(ResponseSpec(
                    name=r_name, power_cfg=pcfg,
                    formula=r_formula_str, weight=r_weight,
                ))
            except (ValueError, TypeError) as e:
                raise ExcelError(f"[RESPONSES] row '{r_name}': {e}") from e

        if len(specs) < 2:
            raise ExcelError(
                f"[RESPONSES] section must define at least 2 responses; found {len(specs)}."
            )
        try:
            multi_cfg = MultiResponseOptions(
                responses=specs, power_combination=power_combination,
            )
        except (ValueError, TypeError) as e:
            raise ExcelError(f"Config sheet produced invalid MultiResponseOptions: {e}") from e

    return formula, factors, power_cfg, design_opts, multi_cfg


# ---------------------------------------------------------------------------
# Output sheet writers
# ---------------------------------------------------------------------------
def _write_df_to_sheet(wb: Any, sheet_name: str, df: pd.DataFrame) -> None:
    """Write *df* to a sheet in *wb*, creating or clearing it as needed."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    _header_row(ws, 1, list(df.columns))
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = ""
            ws.cell(row=r_idx, column=c_idx, value=val)
    _set_column_widths(ws, [max(len(str(c)), 10) + 2 for c in df.columns])


def _write_results_sheet(wb: Any, report: Dict[str, Any]) -> None:
    """Write the run report dict as a key/value table to the Results sheet."""
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")
    _header_row(ws, 1, ["Key", "Value"])
    r = 2
    for key, val in report.items():
        if isinstance(val, dict):  # flatten nested dicts (e.g. diagnostics)
            for sub_key, sub_val in val.items():
                if isinstance(sub_val, list):
                    sub_val = str(sub_val)
                ws.cell(row=r, column=1, value=f"{key}.{sub_key}").font = Font(bold=True)
                ws.cell(row=r, column=2, value=sub_val)
                r += 1
        elif isinstance(val, list):
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=str(val))
            r += 1
        else:
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Template builder
# ---------------------------------------------------------------------------
def create_excel_template(
    path: Union[str, Path] = "iopt_template.xlsx",
    example: str = "r2",
) -> Path:
    """Create a new ``.xlsx`` workbook pre-populated with a runnable example.

    The workbook contains a **Config** sheet that ``excel_run()`` can read
    directly, plus empty placeholder sheets for Results, Design, and Buckets
    that ``excel_run()`` will populate after the design search.

    Dropdown validation is applied to the ``power_mode`` and ``criterion``
    cells.

    Parameters
    ----------
    path : str or Path, default ``"iopt_template.xlsx"``
        Destination file path.  Existing files are overwritten.
    example : {"r2", "contrast"}, default ``"r2"``
        Which pre-filled example to write.

        * ``"r2"``       — global R² power config with two continuous factors.
        * ``"contrast"`` — single contrast (L matrix + delta) with two
          continuous factors.

    Returns
    -------
    Path
        Absolute path of the created workbook.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    ValueError
        If *example* is not ``"r2"`` or ``"contrast"``.
    """
    _require_openpyxl()

    if example not in ("r2", "contrast"):
        raise ValueError(
            f"Unknown example {example!r}. Supported values: 'r2', 'contrast'."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Config"

    r = 1  # current row counter

    # ------------------------------------------------------------------ #
    # [SETTINGS]
    # ------------------------------------------------------------------ #
    _sentinel_style(ws, r, _SENTINEL_SETTINGS); r += 1

    _key_cell(ws, r, "formula",
              "~ 1 + A + B" if example == "r2" else "~ 1 + A + B"); r += 1
    _key_cell(ws, r, "power_mode", example); r += 1
    _add_dropdown(ws, '"r2,contrast"', row=r - 1)

    _key_cell(ws, r, "alpha",        0.05);  r += 1
    _key_cell(ws, r, "power",        0.80);  r += 1

    if example == "r2":
        _key_cell(ws, r, "r2_target",  0.25); r += 1
        _key_cell(ws, r, "sigma",      "");   r += 1   # not used in R² mode
    else:
        _key_cell(ws, r, "sigma",      1.0);  r += 1
        _key_cell(ws, r, "r2_target",  "");   r += 1   # not used in contrast mode

    _key_cell(ws, r, "max_n",        500);   r += 1
    _key_cell(ws, r, "criterion",    "I");   r += 1
    _add_dropdown(ws, '"I,D,A"', row=r - 1)

    _key_cell(ws, r, "starts",       5);     r += 1
    _key_cell(ws, r, "max_iter",     1000);  r += 1
    _key_cell(ws, r, "random_state", 123);   r += 1
    # Blocked design: set n_blocks >= 2 to enable; 0 = unblocked (default).
    _key_cell(ws, r, "n_blocks",                0);       r += 1
    _key_cell(ws, r, "block_factor_name",       "Block"); r += 1
    # Categorical pre-allocation: set to true/false.
    _key_cell(ws, r, "preallocate_categorical", "false"); r += 1
    _add_dropdown(ws, '"true,false"', row=r - 1)
    _key_cell(ws, r, "alloc_min_per_cell",      1);       r += 1
    _key_cell(ws, r, "alloc_max_per_cell",      0);       r += 1  # 0 = no upper limit
    # Split-plot: set htc_factors and n_whole_plots >= 2 to enable.
    _key_cell(ws, r, "htc_factors",             "");      r += 1  # comma-separated HTC factor names
    _key_cell(ws, r, "n_whole_plots",           0);       r += 1  # 0 = disabled; >= 2 to enable
    _key_cell(ws, r, "eta",                     1.0);     r += 1  # variance ratio sigma2_wp/sigma2_sp
    _key_cell(ws, r, "subplots_per_wp",         0);       r += 1  # 0 = auto
    _key_cell(ws, r, "df_method",               "auto");  r += 1
    _add_dropdown(ws, '"auto,conservative,sp_only"', row=r - 1)

    r += 1  # blank separator

    # ------------------------------------------------------------------ #
    # [CONTRAST] — only for contrast example
    # ------------------------------------------------------------------ #
    if example == "contrast":
        _sentinel_style(ws, r, _SENTINEL_CONTRAST); r += 1
        # Contrast between A=-1 and A=+1, holding B at mid-range
        # Model: Intercept + A + B  → L = [0, 1, 0]
        _key_cell(ws, r, "L_row",  "0, 1, 0"); r += 1
        _key_cell(ws, r, "delta",  "1.0");      r += 1
        r += 1  # blank separator

    # ------------------------------------------------------------------ #
    # [FACTORS]
    # ------------------------------------------------------------------ #
    _sentinel_style(ws, r, _SENTINEL_FACTORS); r += 1
    _header_row(ws, r, ["Name", "Type", "Value 1", "Value 2"]); r += 1
    ws.cell(row=r, column=1, value="A")
    ws.cell(row=r, column=2, value="continuous")
    ws.cell(row=r, column=3, value=-1.0)
    ws.cell(row=r, column=4, value=1.0)
    r += 1
    ws.cell(row=r, column=1, value="B")
    ws.cell(row=r, column=2, value="continuous")
    ws.cell(row=r, column=3, value=-1.0)
    ws.cell(row=r, column=4, value=1.0)
    r += 1

    _set_column_widths(ws, [20, 16, 12, 12])

    # ------------------------------------------------------------------ #
    # Placeholder output sheets
    # ------------------------------------------------------------------ #
    for sheet in ("Results", "Design", "Buckets"):
        out_ws = wb.create_sheet(sheet)
        out_ws.cell(row=1, column=1, value=f"Run excel_run() to populate this sheet.")
        out_ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    dest = Path(path).resolve()
    wb.save(dest)
    return dest


# ---------------------------------------------------------------------------
# Run function
# ---------------------------------------------------------------------------
def excel_run(
    path: Union[str, Path],
    design_opts_override: Optional[DesignOptions] = None,
) -> Dict[str, Any]:
    """Read a Config sheet from an ``.xlsx`` workbook, run the design search,
    and write Results / Design / Buckets sheets back into the **same** file.

    Parameters
    ----------
    path : str or Path
        Path to the ``.xlsx`` workbook (must contain a ``Config`` sheet).
    design_opts_override : DesignOptions, optional
        If provided, these options override the values read from the Config
        sheet.  Useful for programmatic control without editing the file.

    Returns
    -------
    dict
        The full result dict returned by ``i_optimal_powered_design``, with
        an additional ``"excel_path"`` key containing the absolute workbook
        path.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    ExcelError
        If the workbook cannot be opened, the Config sheet is missing or
        malformed, or the design run fails.
    """
    _require_openpyxl()
    from .api import i_optimal_powered_design

    path = Path(path).resolve()
    if not path.exists():
        raise ExcelError(f"Workbook not found: {path}")

    # ------------------------------------------------------------------
    # 1. Open and parse Config sheet
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        raise ExcelError(f"Could not open workbook {path}: {e}") from e

    if "Config" not in wb.sheetnames:
        raise ExcelError(
            f"Workbook {path.name!r} has no 'Config' sheet. "
            "Use create_excel_template() to create a properly structured workbook."
        )

    try:
        formula, factors, power_cfg, design_opts, multi_cfg = _read_config_sheet(wb["Config"])
    except ExcelError:
        raise
    except Exception as e:
        raise ExcelError(f"Failed to parse Config sheet: {e}") from e

    if design_opts_override is not None:
        design_opts = design_opts_override

    # ------------------------------------------------------------------
    # 2. Run design search
    # ------------------------------------------------------------------
    try:
        if multi_cfg is not None:
            from .api import i_optimal_multiresponse_design  # noqa: PLC0415
            result = i_optimal_multiresponse_design(
                formula=formula,
                factors=factors,
                multi_cfg=multi_cfg,
                design_opts=design_opts,
            )
        else:
            result = i_optimal_powered_design(
                formula=formula,
                factors=factors,
                power_cfg=power_cfg,
                design_opts=design_opts,
            )
    except Exception as e:
        raise ExcelError(f"Design search failed: {e}") from e

    design_df: pd.DataFrame = result["design_df"]
    buckets_df: pd.DataFrame = result["buckets_df"]
    report: Dict[str, Any] = result["report"]

    # ------------------------------------------------------------------
    # 3. Write output sheets back into the workbook
    # ------------------------------------------------------------------
    try:
        _write_results_sheet(wb, report)
        _write_df_to_sheet(wb, "Design", design_df)
        _write_df_to_sheet(wb, "Buckets", buckets_df)
        wb.save(path)
    except Exception as e:
        raise ExcelError(f"Failed to write results to workbook: {e}") from e

    result["excel_path"] = str(path)
    return result


# ---------------------------------------------------------------------------
# Public exports
# ---------------------------------------------------------------------------
__all__ = ["ExcelError", "create_excel_template", "excel_run"]
