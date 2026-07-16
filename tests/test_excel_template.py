# tests/test_excel_template.py
# License: MIT
"""Unit tests for lattice_doe.excel_template — openpyxl calls mocked where possible."""
from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any, Dict, List
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest

import lattice_doe.excel_template as excel_module
from lattice_doe.excel_template import (
    ExcelError,
    _read_config_sheet,
    create_excel_template,
    excel_run,
)
from lattice_doe.config import PowerContrastConfig, PowerR2Config


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_ws(rows: List[List[Any]]) -> MagicMock:
    """Mock openpyxl Worksheet whose iter_rows(values_only=True) returns *rows*."""
    ws = MagicMock()
    ws.iter_rows.return_value = [tuple(r) for r in rows]
    return ws


def _minimal_result(
    design_df: pd.DataFrame | None = None,
    buckets_df: pd.DataFrame | None = None,
) -> Dict[str, Any]:
    """Minimal result dict matching find_optimal_design() output."""
    if design_df is None:
        design_df = pd.DataFrame({"x1": [0.1, 0.5], "x2": [-1.0, 1.0]})
    if buckets_df is None:
        buckets_df = pd.DataFrame({"x1": [0.1], "count": [2]})
    report = {
        "n": 10,
        "p": 3,
        "df_num": 2,
        "df_denom": 7,
        "alpha": 0.05,
        "target_power": 0.80,
        "achieved_power": 0.85,
        "noncentrality_lambda": 8.0,
        "criterion": "I",
        "elapsed_sec": 1.23,
        "warnings": [],
        "diagnostics": {
            "i_criterion": 0.45,
            "d_efficiency": 0.91,
            "condition_number": 3.2,
        },
    }
    return {"design_df": design_df, "buckets_df": buckets_df, "report": report}


# ---------------------------------------------------------------------------
# TestImportGuard
# ---------------------------------------------------------------------------


# Tests that write a real workbook need openpyxl (the [extras] group); the
# parse-only tests in the same classes run fine on a core install.
requires_openpyxl = pytest.mark.skipif(
    not excel_module._HAS_OPENPYXL, reason="openpyxl not installed"
)

class TestImportGuard:
    def test_create_template_raises_import_error_when_no_openpyxl(self):
        with patch.object(excel_module, "_HAS_OPENPYXL", False):
            with pytest.raises(ImportError, match="openpyxl"):
                create_excel_template("dummy.xlsx")

    def test_excel_run_raises_import_error_when_no_openpyxl(self):
        with patch.object(excel_module, "_HAS_OPENPYXL", False):
            with pytest.raises(ImportError, match="openpyxl"):
                excel_run("dummy.xlsx")


# ---------------------------------------------------------------------------
# TestReadConfigSheet
# ---------------------------------------------------------------------------

class TestReadConfigSheet:
    def test_r2_mode_parses_formula_and_factors(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "r2"],
            ["r2_target", "0.30"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["x2", "continuous", "-1.0", "1.0"],
        ]
        formula, factors, power_cfg, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert formula == "x1 + x2"
        assert factors["x1"] == (-1.0, 1.0)
        assert factors["x2"] == (-1.0, 1.0)

    def test_r2_mode_returns_power_r2_config(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "r2"],
            ["r2_target", "0.25"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        _, _, power_cfg, _, _ = _read_config_sheet(_make_ws(rows))
        assert isinstance(power_cfg, PowerR2Config)
        assert power_cfg.r2_target == pytest.approx(0.25)

    def test_r2_defaults_applied_when_optional_keys_absent(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "r2"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        _, _, power_cfg, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert isinstance(power_cfg, PowerR2Config)
        assert power_cfg.alpha == pytest.approx(0.05)
        assert power_cfg.power == pytest.approx(0.80)
        assert design_opts.starts == 5

    def test_contrast_mode_single_L_row(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "contrast"],
            ["[CONTRAST]", ""],
            ["L_row", "0,1,0"],
            ["delta", "1.0"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["x2", "continuous", "-1.0", "1.0"],
        ]
        _, _, power_cfg, _, _ = _read_config_sheet(_make_ws(rows))
        assert isinstance(power_cfg, PowerContrastConfig)
        assert power_cfg.L.shape == (1, 3)
        np.testing.assert_array_equal(power_cfg.L[0], [0.0, 1.0, 0.0])
        np.testing.assert_array_equal(power_cfg.delta, [1.0])

    def test_contrast_mode_multi_row_L_matrix(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "contrast"],
            ["[CONTRAST]", ""],
            ["L_row", "0,1,0"],
            ["L_row", "0,0,1"],
            ["delta", "1.0,0.5"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["x2", "continuous", "-1.0", "1.0"],
        ]
        _, _, power_cfg, _, _ = _read_config_sheet(_make_ws(rows))
        assert isinstance(power_cfg, PowerContrastConfig)
        assert power_cfg.L.shape == (2, 3)
        np.testing.assert_array_equal(power_cfg.delta, [1.0, 0.5])

    def test_contrast_delta_length_mismatch_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "contrast"],
            ["[CONTRAST]", ""],
            ["L_row", "0,1"],
            ["L_row", "1,0"],
            ["delta", "1.0"],   # 1 value but 2 L_rows
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        with pytest.raises(ExcelError, match="delta"):
            _read_config_sheet(_make_ws(rows))

    def test_missing_settings_sentinel_raises(self):
        rows = [
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        with pytest.raises(ExcelError, match=r"\[SETTINGS\]"):
            _read_config_sheet(_make_ws(rows))

    def test_missing_factors_sentinel_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "r2"],
        ]
        with pytest.raises(ExcelError, match=r"\[FACTORS\]"):
            _read_config_sheet(_make_ws(rows))

    def test_missing_formula_key_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["power_mode", "r2"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        with pytest.raises(ExcelError, match="formula"):
            _read_config_sheet(_make_ws(rows))

    def test_unknown_power_mode_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "bayes"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        with pytest.raises(ExcelError, match="power_mode"):
            _read_config_sheet(_make_ws(rows))

    def test_contrast_missing_sentinel_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "contrast"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        with pytest.raises(ExcelError, match=r"\[CONTRAST\]"):
            _read_config_sheet(_make_ws(rows))

    def test_continuous_factor_parses_to_tuple(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "temp"],
            ["power_mode", "r2"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["temp", "continuous", "20.0", "80.0"],
        ]
        _, factors, _, _, _ = _read_config_sheet(_make_ws(rows))
        assert factors["temp"] == (20.0, 80.0)
        assert isinstance(factors["temp"], tuple)

    def test_categorical_factor_parses_to_list(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "material"],
            ["power_mode", "r2"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2", "value3"],
            ["material", "categorical", "Steel", "Aluminum", "Titanium"],
        ]
        _, factors, _, _, _ = _read_config_sheet(_make_ws(rows))
        assert factors["material"] == ["Steel", "Aluminum", "Titanium"]

    def test_unknown_factor_type_raises(self):
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "r2"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1"],
            ["x1", "ordinal", "low"],
        ]
        with pytest.raises(ExcelError, match="ordinal"):
            _read_config_sheet(_make_ws(rows))

    def test_excel_numeric_int_cells_handled(self):
        """Excel often stores integers as floats ('42.0'); _int() must handle this."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "r2"],
            ["starts", "3.0"],       # float string from Excel numeric cell
            ["random_state", "99.0"],
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.starts == 3
        assert design_opts.random_state == 99


# ---------------------------------------------------------------------------
# TestCreateExcelTemplate  (requires openpyxl installed)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not excel_module._HAS_OPENPYXL, reason="openpyxl not installed")
class TestCreateExcelTemplate:
    def test_r2_template_creates_file(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx", example="r2")
            assert p.exists()
            assert p.stat().st_size > 0

    def test_contrast_template_creates_file(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx", example="contrast")
            assert p.exists()
            assert p.stat().st_size > 0

    def test_returns_absolute_path(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx")
            assert p.is_absolute()

    def test_unknown_example_raises_value_error(self):
        with pytest.raises(ValueError, match="Unknown example"):
            create_excel_template("dummy.xlsx", example="pdf")

    def test_r2_template_config_round_trips(self):
        """The r2 template must produce a parseable Config sheet."""
        import openpyxl
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx", example="r2")
            wb = openpyxl.load_workbook(p)
            ws = wb["Config"]
            formula, factors, power_cfg, design_opts, _ = _read_config_sheet(ws)
        assert isinstance(power_cfg, PowerR2Config)
        assert formula  # non-empty
        assert factors   # at least one factor

    def test_contrast_template_config_round_trips(self):
        """The contrast template must produce a parseable Config sheet."""
        import openpyxl
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx", example="contrast")
            wb = openpyxl.load_workbook(p)
            ws = wb["Config"]
            formula, factors, power_cfg, design_opts, _ = _read_config_sheet(ws)
        assert isinstance(power_cfg, PowerContrastConfig)

    def test_workbook_has_expected_sheets(self):
        """Workbook should contain Config, Results, Design, Buckets placeholders."""
        import openpyxl
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test.xlsx", example="r2")
            wb = openpyxl.load_workbook(p)
        assert "Config" in wb.sheetnames


# ---------------------------------------------------------------------------
# TestExcelRun  (requires openpyxl installed)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not excel_module._HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelRun:
    def test_r2_run_returns_expected_keys(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_r2.xlsx", example="r2")
            result = excel_run(p)
        assert "design_df" in result
        assert "buckets_df" in result
        assert "report" in result
        assert "excel_path" in result

    def test_r2_run_design_df_is_dataframe(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_r2.xlsx", example="r2")
            result = excel_run(p)
        assert isinstance(result["design_df"], pd.DataFrame)
        assert len(result["design_df"]) > 0

    def test_r2_run_excel_path_is_same_file(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_r2.xlsx", example="r2")
            result = excel_run(p)
        assert Path(result["excel_path"]).resolve() == p.resolve()

    def test_r2_run_writes_output_sheets(self):
        """After excel_run, the workbook must contain Results, Design, Buckets sheets."""
        import openpyxl
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_r2.xlsx", example="r2")
            excel_run(p)
            wb = openpyxl.load_workbook(p)
        for expected in ("Results", "Design", "Buckets"):
            assert expected in wb.sheetnames, f"Missing sheet: {expected}"

    def test_contrast_run_returns_power_contrast_config_type(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_c.xlsx", example="contrast")
            result = excel_run(p)
        r = result["report"]
        assert r["achieved_power"] > 0

    def test_design_failure_raises_excel_error(self):
        with tempfile.TemporaryDirectory() as td:
            p = create_excel_template(Path(td) / "test_r2.xlsx", example="r2")
            with patch(
                "lattice_doe.api.find_optimal_design",
                side_effect=RuntimeError("solver failed"),
            ):
                with pytest.raises(ExcelError, match="Design search failed"):
                    excel_run(p)

    def test_missing_file_raises_excel_error(self):
        with pytest.raises(ExcelError, match="not found"):
            excel_run("/nonexistent/path/file.xlsx")


# ---------------------------------------------------------------------------
# CR-22: blocked/pre-allocation fields in Excel parser and template
# ---------------------------------------------------------------------------

class TestCR22BlockedPreAllocFields:
    """Verify n_blocks, block_factor_name, preallocate_categorical,
    alloc_min_per_cell, and alloc_max_per_cell are parsed and propagated."""

    def _base_rows(self, extra_settings):
        return [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "r2"],
            ["r2_target", "0.25"],
        ] + extra_settings + [
            ["[FACTORS]", ""],
            ["factor_name", "type", "value1", "value2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]

    def test_n_blocks_zero_leaves_unblocked(self):
        rows = self._base_rows([["n_blocks", "0"]])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.n_blocks is None

    def test_n_blocks_2_enables_blocking(self):
        rows = self._base_rows([
            ["n_blocks", "2"],
            ["block_factor_name", "Batch"],
        ])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.n_blocks == 2
        assert design_opts.block_factor_name == "Batch"

    def test_block_factor_name_default_is_block(self):
        rows = self._base_rows([["n_blocks", "3"]])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.block_factor_name == "Block"

    def test_preallocate_categorical_true(self):
        rows = self._base_rows([
            ["preallocate_categorical", "true"],
            ["alloc_min_per_cell", "2"],
        ])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.preallocate_categorical is True
        assert design_opts.alloc_min_per_cell == 2

    def test_preallocate_categorical_false_by_default(self):
        rows = self._base_rows([])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.preallocate_categorical is False

    def test_alloc_max_per_cell_zero_maps_to_none(self):
        rows = self._base_rows([
            ["preallocate_categorical", "true"],
            ["alloc_max_per_cell", "0"],
        ])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.alloc_max_per_cell is None

    def test_alloc_max_per_cell_positive_is_forwarded(self):
        rows = self._base_rows([
            ["preallocate_categorical", "true"],
            ["alloc_max_per_cell", "5"],
        ])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.alloc_max_per_cell == 5

    def test_excel_numeric_cell_for_n_blocks(self):
        """Excel numeric cells arrive as '2.0' — int(float()) must handle it."""
        rows = self._base_rows([["n_blocks", "2.0"]])
        _, _, _, design_opts, _ = _read_config_sheet(_make_ws(rows))
        assert design_opts.n_blocks == 2

    def test_invalid_bool_raises(self):
        rows = self._base_rows([["preallocate_categorical", "maybe"]])
        with pytest.raises(ExcelError, match="true/false"):
            _read_config_sheet(_make_ws(rows))

    @requires_openpyxl
    def test_template_contains_new_keys(self):
        """create_excel_template() must write all 5 new setting rows."""
        with tempfile.TemporaryDirectory() as tmp:
            for example in ("r2", "contrast"):
                p = Path(tmp) / f"tpl_{example}.xlsx"
                create_excel_template(p, example=example)
                import openpyxl
                wb = openpyxl.load_workbook(p)
                ws = wb["Config"]
                col_a_values = [
                    ws.cell(row=r, column=1).value
                    for r in range(1, ws.max_row + 1)
                ]
                for key in (
                    "n_blocks", "block_factor_name",
                    "preallocate_categorical", "alloc_min_per_cell", "alloc_max_per_cell",
                ):
                    assert key in col_a_values, (
                        f"{example!r} template missing '{key}' in Config sheet"
                    )

    @requires_openpyxl
    def test_template_still_parseable(self):
        """All templates must round-trip through _read_config_sheet cleanly."""
        with tempfile.TemporaryDirectory() as tmp:
            for example in ("r2", "contrast", "multiresponse"):
                p = Path(tmp) / f"tpl_{example}.xlsx"
                create_excel_template(p, example=example)
                import openpyxl
                wb = openpyxl.load_workbook(p)
                formula, factors, power_cfg, design_opts, multi_cfg = _read_config_sheet(wb["Config"])
                assert formula
                assert factors
                if example == "multiresponse":
                    assert multi_cfg is not None
                else:
                    assert power_cfg is not None


# ---------------------------------------------------------------------------
# TestCR34ExcelMultiResponseAdvancedFields
# ---------------------------------------------------------------------------

class TestCR34ExcelMultiResponseAdvancedFields:
    """CR-34: [RESPONSES] parser must accept sigma_joint and advanced per-response knobs."""

    def _base_rows(self, response_rows: list) -> list:
        return [
            ("[SETTINGS]", None),
            ("formula", "x1 + x2"),
            ("[FACTORS]", None),
            ("factor_name", "type", "value1", "value2"),
            ("x1", "continuous", -1.0, 1.0),
            ("x2", "continuous", -1.0, 1.0),
            ("[RESPONSES]", None),
            ("name", "power_mode", "sigma", "alpha", "power", "weight",
             "L_row", "delta", "r2_target", "formula",
             "lambda_mode", "max_n", "max_iter", "tol_power"),
        ] + [tuple(r) if isinstance(r, list) else r for r in response_rows]

    def _ws(self, response_rows: list):
        return _make_ws(self._base_rows(response_rows))

    def test_lambda_mode_forwarded(self):
        ws = self._ws([
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, "n_minus_p", None, None, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, "n", None, None, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.responses[0].power_cfg.lambda_mode == "n_minus_p"
        assert multi_cfg.responses[1].power_cfg.lambda_mode == "n"

    def test_max_n_forwarded(self):
        ws = self._ws([
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, 300, None, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, 400, None, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.responses[0].power_cfg.max_n == 300
        assert multi_cfg.responses[1].power_cfg.max_n == 400

    def test_max_iter_forwarded(self):
        ws = self._ws([
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, None, 50, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, 100, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.responses[0].power_cfg.max_iter == 50

    def test_tol_power_forwarded(self):
        ws = self._ws([
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, None, None, 0.005),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, None, 0.002),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.responses[0].power_cfg.tol_power == pytest.approx(0.005)

    def test_contrast_response_advanced_knobs(self):
        ws = self._ws([
            ("Y1", "contrast", 1.0, None, None, 1.0, "0,1,0", "1.0", None, None, None, 200, 50, 0.002),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, None, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert isinstance(multi_cfg.responses[0].power_cfg, PowerContrastConfig)
        assert multi_cfg.responses[0].power_cfg.max_n == 200
        assert multi_cfg.responses[0].power_cfg.tol_power == pytest.approx(0.002)

    def test_sigma_joint_parsed_2x2(self):
        ws = self._ws([
            ("power_combination", "min"),
            ("sigma_joint", "1.0,0.3; 0.3,1.0"),
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, None, None, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, None, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.sigma_joint is not None
        assert multi_cfg.sigma_joint.shape == (2, 2)
        assert multi_cfg.sigma_joint[0, 1] == pytest.approx(0.3)

    def test_sigma_joint_none_gives_none(self):
        ws = self._ws([
            ("sigma_joint", None),
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, None, None, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, None, None),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.sigma_joint is None

    def test_sigma_joint_invalid_raises(self):
        ws = self._ws([
            ("sigma_joint", "1.0,abc; 0.3,1.0"),
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15, None, None, None, None, None),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20, None, None, None, None, None),
        ])
        with pytest.raises(ExcelError, match="sigma_joint"):
            _read_config_sheet(ws)

    def test_defaults_when_advanced_cols_absent(self):
        ws = self._ws([
            ("Y1", "r2", None, None, None, 1.0, None, None, 0.15),
            ("Y2", "r2", None, None, None, 1.0, None, None, 0.20),
        ])
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        assert multi_cfg.responses[0].power_cfg.lambda_mode == "n"
        assert multi_cfg.responses[0].power_cfg.max_n == 2000
        assert multi_cfg.responses[0].power_cfg.max_iter == 200
        assert multi_cfg.responses[0].power_cfg.tol_power == pytest.approx(1e-3)

    @requires_openpyxl
    def test_multiresponse_template_round_trips(self):
        """The multiresponse Excel template must parse correctly."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "mr_template.xlsx"
            create_excel_template(p, example="multiresponse")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            _, _, power_cfg, _, multi_cfg = _read_config_sheet(wb["Config"])
            assert power_cfg is None
            assert multi_cfg is not None
            assert len(multi_cfg.responses) == 2

    @requires_openpyxl
    def test_multiresponse_template_sigma_joint_none(self):
        """Built-in multiresponse template has blank sigma_joint → None."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "mr_template.xlsx"
            create_excel_template(p, example="multiresponse")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            _, _, _, _, multi_cfg = _read_config_sheet(wb["Config"])
            assert multi_cfg.sigma_joint is None

    @requires_openpyxl
    def test_unknown_multiresponse_example_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            with pytest.raises(ValueError, match="bayesian"):
                create_excel_template(Path(tmp) / "x.xlsx", example="bayesian")


# ---------------------------------------------------------------------------
# TestExcelGLMSupport — GL-9: GLM support in Excel connector
# ---------------------------------------------------------------------------

class TestExcelGLMSupport:
    """Tests for GLM power mode in the Excel connector (GL-9)."""

    def test_excel_glm_settings_parse(self):
        """power_mode='glm' with family/baseline parses correctly."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["power_mode", "glm"],
            ["family", "binomial"],
            ["link", ""],
            ["baseline", "0.20"],
            ["[CONTRAST]", ""],
            ["L_row", "0,1,0"],
            ["delta", "0.5"],
            ["[FACTORS]", ""],
            ["Name", "Type", "Value 1", "Value 2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["x2", "continuous", "-1.0", "1.0"],
        ]
        ws = _make_ws(rows)
        formula, factors, power_cfg, design_opts, multi_cfg = _read_config_sheet(ws)
        from lattice_doe.config import PowerGLMContrastConfig
        assert isinstance(power_cfg, PowerGLMContrastConfig)
        assert formula == "x1 + x2"
        assert multi_cfg is None

    def test_excel_glm_family_builds_glm_config(self):
        """GLM config attributes (family, baseline, L, delta) are forwarded correctly."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "glm"],
            ["family", "poisson"],
            ["baseline", "2.5"],
            ["[CONTRAST]", ""],
            ["L_row", "0,1"],
            ["delta", "0.4"],
            ["[FACTORS]", ""],
            ["Name", "Type", "Value 1", "Value 2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        ws = _make_ws(rows)
        _, _, power_cfg, _, _ = _read_config_sheet(ws)
        from lattice_doe.config import PowerGLMContrastConfig
        assert isinstance(power_cfg, PowerGLMContrastConfig)
        assert power_cfg.family == "poisson"
        assert power_cfg.baseline == pytest.approx(2.5)
        np.testing.assert_array_equal(power_cfg.delta, [0.4])

    def test_excel_glm_missing_baseline_raises(self):
        """Omitting baseline when power_mode='glm' raises ExcelError."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["power_mode", "glm"],
            ["family", "binomial"],
            # baseline omitted
            ["[CONTRAST]", ""],
            ["L_row", "0,1"],
            ["delta", "0.5"],
            ["[FACTORS]", ""],
            ["Name", "Type", "Value 1", "Value 2"],
            ["x1", "continuous", "-1.0", "1.0"],
        ]
        ws = _make_ws(rows)
        with pytest.raises(ExcelError, match="baseline"):
            _read_config_sheet(ws)

    @requires_openpyxl
    def test_excel_glm_template_binomial_creates(self):
        """create_excel_template('glm-binomial') creates a valid .xlsx file."""
        with tempfile.TemporaryDirectory() as tmp:
            p = create_excel_template(Path(tmp) / "glm_bin.xlsx", example="glm-binomial")
            assert p.exists()
            assert p.suffix == ".xlsx"

    @requires_openpyxl
    def test_excel_glm_template_round_trips(self):
        """'glm-binomial' template can be written and parsed back."""
        with tempfile.TemporaryDirectory() as tmp:
            p = create_excel_template(Path(tmp) / "glm_bin.xlsx", example="glm-binomial")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            _, _, power_cfg, _, multi_cfg = _read_config_sheet(wb["Config"])
            from lattice_doe.config import PowerGLMContrastConfig
            assert isinstance(power_cfg, PowerGLMContrastConfig)
            assert power_cfg.family == "binomial"
            assert multi_cfg is None

    def test_excel_responses_glm_family_per_row(self):
        """[RESPONSES] row with power_mode='glm' builds PowerGLMContrastConfig."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1 + x2"],
            ["[FACTORS]", ""],
            ["Name", "Type", "Value 1", "Value 2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["x2", "continuous", "-1.0", "1.0"],
            ["[RESPONSES]", ""],
            ["name", "power_mode", "sigma", "alpha", "power", "weight",
             "L_row", "delta", "r2_target", "formula", "lambda_mode",
             "max_n", "max_iter", "tol_power", "family", "baseline"],
            ["power_combination", "min"],
            ["sigma_joint", ""],
            ["Y1", "glm", "", "", "", "1.0",
             "0,1,0", "0.5", "", "", "", "", "", "",
             "binomial", "0.20"],
            ["Y2", "r2", "", "", "", "1.0",
             "", "", "0.15", "", "n", "", "", ""],
        ]
        ws = _make_ws(rows)
        _, _, power_cfg, _, multi_cfg = _read_config_sheet(ws)
        assert power_cfg is None
        assert multi_cfg is not None
        from lattice_doe.config import PowerGLMContrastConfig, PowerR2Config
        assert isinstance(multi_cfg.responses[0].power_cfg, PowerGLMContrastConfig)
        assert isinstance(multi_cfg.responses[1].power_cfg, PowerR2Config)

    def test_excel_responses_glm_baseline_forwarded(self):
        """Baseline from col 16 is forwarded to PowerGLMContrastConfig."""
        rows = [
            ["[SETTINGS]", ""],
            ["formula", "x1"],
            ["[FACTORS]", ""],
            ["Name", "Type", "Value 1", "Value 2"],
            ["x1", "continuous", "-1.0", "1.0"],
            ["[RESPONSES]", ""],
            ["name", "power_mode", "sigma", "alpha", "power", "weight",
             "L_row", "delta", "r2_target", "formula", "lambda_mode",
             "max_n", "max_iter", "tol_power", "family", "baseline"],
            ["power_combination", "min"],
            ["sigma_joint", ""],
            ["Y1", "glm", "", "", "", "1.0",
             "0,1", "0.4", "", "", "", "", "", "",
             "poisson", "3.5"],
            ["Y2", "r2", "", "", "", "1.0",
             "", "", "0.20", "", "n", "", "", ""],
        ]
        ws = _make_ws(rows)
        _, _, _, _, multi_cfg = _read_config_sheet(ws)
        from lattice_doe.config import PowerGLMContrastConfig
        y1 = multi_cfg.responses[0].power_cfg
        assert isinstance(y1, PowerGLMContrastConfig)
        assert y1.baseline == pytest.approx(3.5)

    @requires_openpyxl
    def test_linear_excel_unchanged(self):
        """Existing r2 template still round-trips correctly (regression guard)."""
        with tempfile.TemporaryDirectory() as tmp:
            p = create_excel_template(Path(tmp) / "r2.xlsx", example="r2")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            _, _, power_cfg, _, _ = _read_config_sheet(wb["Config"])
            assert isinstance(power_cfg, PowerR2Config)

    @requires_openpyxl
    def test_glm_template_poisson_creates_and_parses(self):
        """'glm-poisson' template can be created and parsed back."""
        with tempfile.TemporaryDirectory() as tmp:
            p = create_excel_template(Path(tmp) / "glm_poi.xlsx", example="glm-poisson")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            _, _, power_cfg, _, _ = _read_config_sheet(wb["Config"])
            from lattice_doe.config import PowerGLMContrastConfig
            assert isinstance(power_cfg, PowerGLMContrastConfig)
            assert power_cfg.family == "poisson"

    @requires_openpyxl
    def test_glm_template_has_baseline_row(self):
        """'glm-binomial' template Config sheet contains a 'baseline' key."""
        with tempfile.TemporaryDirectory() as tmp:
            p = create_excel_template(Path(tmp) / "glm_b.xlsx", example="glm-binomial")
            import openpyxl
            wb = openpyxl.load_workbook(p)
            ws = wb["Config"]
            col_a_values = [
                str(ws.cell(row=r, column=1).value or "").strip()
                for r in range(1, ws.max_row + 1)
            ]
            assert "baseline" in col_a_values


class TestCompoundMatrixExcelExport:
    """UX-66: the Excel writer must export every compound per-response
    matrix as a safely named sheet plus a name-to-sheet index (sheet titles
    cap at 31 chars and forbid path separators)."""

    @staticmethod
    def _compound_result():
        design_df = pd.DataFrame({"x1": [0.1, 0.5]})
        return {
            "design_df": design_df,
            "buckets_df": pd.DataFrame({"x1": [0.1], "count": [2]}),
            "report": {"n": 2, "compound_criterion": True,
                       "achieved_power": 0.5, "warnings": []},
            "model_matrix": pd.DataFrame({"Intercept": [1.0, 1.0],
                                          "x1": [0.1, 0.5]}),
            "model_matrices": {
                "y1": pd.DataFrame({"Intercept": [1.0, 1.0],
                                    "x1": [0.1, 0.5]}),
                "Yield/Day": pd.DataFrame({"Intercept": [1.0, 1.0],
                                           "bs(x1)[0]": [0.2, 0.3]}),
            },
        }

    def test_per_response_sheets_and_index(self):
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])

        for expected in ("ModelMatrix", "MM_y1", "MM_Yield_Day",
                         "ModelMatrixIndex"):
            assert expected in wb.sheetnames, wb.sheetnames

        idx = wb["ModelMatrixIndex"]
        rows = [tuple(c.value for c in r) for r in idx.iter_rows()]
        assert ("response", "sheet") == rows[0]
        assert ("Yield/Day", "MM_Yield_Day") in rows   # original name kept

        mm2 = wb["MM_Yield_Day"]
        header = [c.value for c in next(mm2.iter_rows())]
        assert header == ["Intercept", "bs(x1)[0]"]

    def test_non_compound_writes_no_extra_sheets(self):
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        res["report"]["compound_criterion"] = False
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        assert "ModelMatrixIndex" not in wb.sheetnames
        assert not any(n.startswith("MM_") for n in wb.sheetnames)

    def test_case_only_response_names_keep_index_consistent(self):
        """UX-69: openpyxl silently renames a case-colliding sheet
        (MM_yield -> MM_yield1), leaving the index pointing at a sheet that
        does not exist. Slugs must diverge before openpyxl ever sees them,
        and every index entry must reference a real sheet."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        res["model_matrices"] = {
            "Yield": pd.DataFrame({"Intercept": [1.0, 1.0]}),
            "yield": pd.DataFrame({"Intercept": [1.0, 1.0]}),
        }
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])

        assert "MM_Yield" in wb.sheetnames
        assert "MM_yield_2" in wb.sheetnames
        idx_rows = [tuple(c.value for c in r)
                    for r in wb["ModelMatrixIndex"].iter_rows()][1:]
        assert ("yield", "MM_yield_2") in idx_rows
        # the desync itself: every index entry must point at a real sheet
        assert all(sheet in wb.sheetnames for _, sheet in idx_rows)

    @staticmethod
    def _idx_rows(wb):
        return [tuple(c.value for c in r)
                for r in wb["ModelMatrixIndex"].iter_rows(min_row=2)]

    def test_repeat_export_with_case_changed_name_replaces_sheets(self):
        """UX-73: exporting again into the SAME workbook after a response
        was renamed by case only must REPLACE the previous per-response
        sheets. Seeding collision detection with prefixed titles while
        checking the bare slug missed the collision, so openpyxl renamed
        the new sheet (MM_yield1) while the index recorded MM_yield."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        res["model_matrices"] = {
            "Yield": pd.DataFrame({"Intercept": [1.0, 1.0]}),
            "Other": pd.DataFrame({"Intercept": [1.0, 1.0]}),
        }
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        assert "MM_Yield" in wb.sheetnames

        res2 = self._compound_result()
        res2["model_matrices"] = {
            "yield": pd.DataFrame({"Intercept": [2.0, 2.0]}),
            "Other": pd.DataFrame({"Intercept": [1.0, 1.0]}),
        }
        _write_output_sheets(wb, res2, res2["report"],
                             res2["design_df"], res2["buckets_df"])

        mm = sorted(n for n in wb.sheetnames if n.startswith("MM_"))
        assert mm == ["MM_Other", "MM_yield"], mm   # no MM_Yield, no MM_yield1
        idx_rows = self._idx_rows(wb)
        assert idx_rows == [("yield", "MM_yield"), ("Other", "MM_Other")]
        assert all(sheet in wb.sheetnames for _, sheet in idx_rows)
        # the replaced sheet carries the SECOND export's matrix
        assert wb["MM_yield"].cell(row=2, column=1).value == 2.0

    def test_users_own_prefixed_sheet_survives_and_forces_suffix(self):
        """UX-73: only sheets OUR index recorded are reconciled away — a
        user's own MM_-style sheet survives, and the new title must dodge
        it case-insensitively (Excel titles do not distinguish case)."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        wb = openpyxl.Workbook()
        wb.create_sheet("mm_YIELD")  # the user's sheet; never indexed by us

        res = self._compound_result()
        res["model_matrices"] = {
            "Yield": pd.DataFrame({"Intercept": [1.0, 1.0]}),
            "Other": pd.DataFrame({"Intercept": [1.0, 1.0]}),
        }
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])

        assert "mm_YIELD" in wb.sheetnames           # untouched
        idx_rows = self._idx_rows(wb)
        assert ("Yield", "MM_Yield_2") in idx_rows   # dodged the collision
        assert all(sheet in wb.sheetnames for _, sheet in idx_rows)

    def test_edited_index_cannot_delete_user_sheets(self):
        """UX-75: ModelMatrixIndex cells are ordinary user-editable content
        and must never authorize deletion — ownership comes from the
        workbook's custom document property instead."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])

        user = wb.create_sheet("MM_UserNotes")
        user["A1"] = "irreplaceable analysis notes"
        wb["ModelMatrixIndex"].append(("notes", "MM_UserNotes"))  # hand edit

        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        assert "MM_UserNotes" in wb.sheetnames
        assert wb["MM_UserNotes"]["A1"].value == "irreplaceable analysis notes"
        # our own sheets were still legitimately replaced
        assert "MM_y1" in wb.sheetnames

    def test_unrelated_user_index_survives_plain_export(self):
        """UX-75: a user's own sheet that happens to be titled
        ModelMatrixIndex is not ours to delete on a non-compound export."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        wb = openpyxl.Workbook()
        own = wb.create_sheet("ModelMatrixIndex")
        own["A1"] = "my personal index of things"

        res = self._compound_result()
        res["report"]["compound_criterion"] = False
        res["model_matrices"] = None
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        assert "ModelMatrixIndex" in wb.sheetnames
        assert wb["ModelMatrixIndex"]["A1"].value == (
            "my personal index of things"
        )

    def test_ownership_marker_survives_save_and_load(self):
        """UX-75: the custom document property must round-trip through a
        real save/load cycle, or repeat exports into a saved workbook would
        have no deletion authority and could never reconcile."""
        import io

        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        res["model_matrices"] = {"Yield": pd.DataFrame({"Intercept": [1.0]})}
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb2 = openpyxl.load_workbook(buf)

        res2 = self._compound_result()
        res2["model_matrices"] = {"yield": pd.DataFrame({"Intercept": [2.0]})}
        _write_output_sheets(wb2, res2, res2["report"],
                             res2["design_df"], res2["buckets_df"])
        mm = [n for n in wb2.sheetnames if n.startswith("MM_")]
        assert mm == ["MM_yield"], mm      # old sheet reconciled after reload
        assert all(sheet in wb2.sheetnames
                   for _, sheet in self._idx_rows(wb2))

    def test_legacy_pre_marker_workbook_warns_and_leaves_residue(self):
        """UX-79: sheets from a release without the ownership marker can
        never be verified ours, so they are left untouched — permanently —
        and the export says so instead of silently accumulating stale
        siblings next to current ones."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        wb = openpyxl.Workbook()                       # NO ownership marker
        legacy = wb.create_sheet("MM_Yield")
        legacy["A1"] = "Intercept"
        legacy["A2"] = 0.5
        legacy_idx = wb.create_sheet("ModelMatrixIndex")
        legacy_idx.append(("response", "sheet"))
        legacy_idx.append(("Yield", "MM_Yield"))

        res = self._compound_result()
        res["model_matrices"] = {
            "Yield": pd.DataFrame({"Intercept": [9.0]}),
        }
        with pytest.warns(RuntimeWarning,
                          match="predates ownership tracking.*MM_Yield"):
            _write_output_sheets(wb, res, res["report"],
                                 res["design_df"], res["buckets_df"])

        assert wb["MM_Yield"]["A2"].value == 0.5       # legacy untouched
        assert ("Yield", "MM_Yield_2") in self._idx_rows(wb)
        assert wb["MM_Yield_2"]["A2"].value == 9.0

        # index and new sheets are now marker-tracked: the next export
        # replaces current output without warning again
        import warnings as W
        res2 = self._compound_result()
        res2["model_matrices"] = {
            "Yield": pd.DataFrame({"Intercept": [10.0]}),
        }
        with W.catch_warnings():
            W.simplefilter("error", RuntimeWarning)
            _write_output_sheets(wb, res2, res2["report"],
                                 res2["design_df"], res2["buckets_df"])
        assert wb["MM_Yield"]["A2"].value == 0.5
        assert wb["MM_Yield_2"]["A2"].value == 10.0

    def test_non_compound_repeat_export_clears_stale_compound_output(self):
        """UX-73: output sheets must describe THIS result only. After a
        compound export, re-exporting a non-compound result into the same
        workbook must remove the per-response sheets and the index — and a
        result without a model matrix must not keep the old ModelMatrix
        next to the new Design."""
        import openpyxl

        from lattice_doe.excel_template import _write_output_sheets

        res = self._compound_result()
        wb = openpyxl.Workbook()
        _write_output_sheets(wb, res, res["report"],
                             res["design_df"], res["buckets_df"])
        assert any(n.startswith("MM_") for n in wb.sheetnames)

        res2 = self._compound_result()
        res2["report"]["compound_criterion"] = False
        res2["model_matrices"] = None
        _write_output_sheets(wb, res2, res2["report"],
                             res2["design_df"], res2["buckets_df"])
        assert not any(n.startswith("MM_") for n in wb.sheetnames)
        assert "ModelMatrixIndex" not in wb.sheetnames
        assert "ModelMatrix" in wb.sheetnames        # still carried a basis

        res3 = self._compound_result()
        res3["report"]["compound_criterion"] = False
        res3["model_matrices"] = None
        res3["model_matrix"] = None
        _write_output_sheets(wb, res3, res3["report"],
                             res3["design_df"], res3["buckets_df"])
        assert "ModelMatrix" not in wb.sheetnames
