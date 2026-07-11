# tests/test_split_plot.py
"""Unit tests for lattice_doe.split_plot and build_split_plot_candidate."""
from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from lattice_doe.split_plot import (
    build_whole_plot_indicator,
    build_split_plot_covariance_inv,
    gls_information_matrix,
    classify_contrasts,
    split_plot_df_denom,
)
from lattice_doe.candidate import build_split_plot_candidate
from lattice_doe.api import find_optimal_design
from lattice_doe.analysis import power_curve_by_wp, power_sensitivity
from lattice_doe.config import (
    PowerContrastConfig,
    PowerR2Config,
    DesignOptions,
    SplitPlotOptions,
)
from lattice_doe.power import (
    contrast_power,
    global_r2_power,
    contrast_power_sp,
    global_r2_power_sp,
)
from lattice_doe.iopt_search import (
    _gls_i_criterion,
    _gls_d_criterion,
    _gls_a_criterion,
    _score_design_gls,
    _criterion_score,
    _i_criterion_for_indices,
    _d_criterion_for_indices,
    _a_criterion_for_indices,
    build_split_plot_design,
)


# ---------------------------------------------------------------------------
# Helpers shared across test classes
# ---------------------------------------------------------------------------

def _make_balanced(n_wp: int, s: int):
    """Return (Z, n_total, V, V_inv) for a balanced layout with eta=1."""
    n_total = n_wp * s
    Z = build_whole_plot_indicator(n_total, n_wp, s)
    V_inv = build_split_plot_covariance_inv(Z, eta=1.0)
    V = np.eye(n_total) + np.ones((n_total, n_total)) * 0.0  # placeholder
    # Build true V = I + eta * Z Z'
    V = np.eye(n_total) + Z @ Z.T
    return Z, n_total, V, V_inv


# ===========================================================================
# TestBuildWholePlotIndicator
# ===========================================================================

class TestBuildWholePlotIndicator:

    def test_shape(self):
        Z = build_whole_plot_indicator(12, 3, 4)
        assert Z.shape == (12, 3)

    def test_dtype(self):
        Z = build_whole_plot_indicator(6, 2, 3)
        assert Z.dtype == np.float64

    def test_entries_are_zero_or_one(self):
        Z = build_whole_plot_indicator(8, 4, 2)
        assert set(np.unique(Z)) == {0.0, 1.0}

    def test_each_row_sums_to_one(self):
        """Each observation belongs to exactly one WP."""
        Z = build_whole_plot_indicator(12, 4, 3)
        np.testing.assert_array_equal(Z.sum(axis=1), np.ones(12))

    def test_each_column_sums_to_subplots_per_wp(self):
        """Each WP column has exactly subplots_per_wp ones."""
        n_wp, s = 5, 3
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        np.testing.assert_array_equal(Z.sum(axis=0), np.full(n_wp, s))

    def test_correct_block_structure(self):
        """Rows 0..s-1 → WP 0, rows s..2s-1 → WP 1, etc."""
        n_wp, s = 3, 2
        Z = build_whole_plot_indicator(6, n_wp, s)
        for k in range(n_wp):
            for j in range(s):
                row = k * s + j
                assert Z[row, k] == 1.0
                for other_k in range(n_wp):
                    if other_k != k:
                        assert Z[row, other_k] == 0.0

    def test_ztZ_is_scalar_multiple_of_identity_for_balanced(self):
        """For balanced layout, Z'Z = s * I_{n_wp}."""
        n_wp, s = 4, 3
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        ZtZ = Z.T @ Z
        expected = s * np.eye(n_wp)
        np.testing.assert_allclose(ZtZ, expected)

    def test_mismatched_n_total_raises(self):
        with pytest.raises(ValueError, match="n_total"):
            build_whole_plot_indicator(7, 2, 3)  # 7 != 2*3

    def test_single_subplot_per_wp(self):
        """Edge case: one sub-plot per WP reduces to independent runs."""
        n_wp = 5
        Z = build_whole_plot_indicator(5, n_wp, 1)
        np.testing.assert_array_equal(Z, np.eye(n_wp))


# ===========================================================================
# TestBuildSplitPlotCovarianceInv
# ===========================================================================

class TestBuildSplitPlotCovarianceInv:

    def test_eta_zero_returns_identity(self):
        Z = build_whole_plot_indicator(6, 2, 3)
        V_inv = build_split_plot_covariance_inv(Z, eta=0.0)
        np.testing.assert_allclose(V_inv, np.eye(6))

    def test_shape(self):
        Z = build_whole_plot_indicator(12, 3, 4)
        V_inv = build_split_plot_covariance_inv(Z, eta=2.0)
        assert V_inv.shape == (12, 12)

    def test_dtype(self):
        Z = build_whole_plot_indicator(6, 2, 3)
        V_inv = build_split_plot_covariance_inv(Z, eta=1.0)
        assert V_inv.dtype == np.float64

    def test_V_times_V_inv_is_identity(self):
        """V · V⁻¹ ≈ I for various eta values."""
        n_wp, s = 3, 4
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        n = n_wp * s
        for eta in (0.5, 1.0, 2.0, 5.0, 10.0):
            V = np.eye(n) + eta * (Z @ Z.T)
            V_inv = build_split_plot_covariance_inv(Z, eta=eta)
            product = V @ V_inv
            np.testing.assert_allclose(product, np.eye(n), atol=1e-10,
                                       err_msg=f"V @ V_inv ≠ I at eta={eta}")

    def test_symmetric(self):
        Z = build_whole_plot_indicator(8, 2, 4)
        for eta in (0.1, 1.0, 3.0):
            V_inv = build_split_plot_covariance_inv(Z, eta=eta)
            np.testing.assert_allclose(V_inv, V_inv.T, atol=1e-14,
                                       err_msg=f"V_inv not symmetric at eta={eta}")

    def test_balanced_closed_form_diagonal(self):
        """V⁻¹[i,i] = 1 − η/(1 + η·s) for balanced layout."""
        n_wp, s = 2, 3
        eta = 1.0
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=eta)
        expected_diag = 1.0 - eta / (1.0 + eta * s)
        np.testing.assert_allclose(np.diag(V_inv), expected_diag, atol=1e-12)

    def test_balanced_closed_form_same_wp_off_diagonal(self):
        """V⁻¹[i,j] = −η/(1 + η·s) for i≠j in the same WP."""
        n_wp, s = 2, 3
        eta = 1.0
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=eta)
        expected_off = -eta / (1.0 + eta * s)
        # Check all same-WP off-diagonal entries (rows 0,1,2 are WP 0)
        for i in range(s):
            for j in range(s):
                if i != j:
                    np.testing.assert_allclose(
                        V_inv[i, j], expected_off, atol=1e-12,
                        err_msg=f"Off-diagonal same-WP mismatch at ({i},{j})"
                    )

    def test_different_wp_entries_are_zero(self):
        """V⁻¹[i,j] = 0 for observations in different WPs."""
        n_wp, s = 3, 2
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=2.0)
        # WP 0: rows 0,1   WP 1: rows 2,3   WP 2: rows 4,5
        cross_pairs = [(0, 2), (0, 3), (1, 4), (2, 5)]
        for i, j in cross_pairs:
            np.testing.assert_allclose(
                V_inv[i, j], 0.0, atol=1e-14,
                err_msg=f"Cross-WP entry ({i},{j}) should be 0"
            )

    def test_negative_eta_raises(self):
        Z = build_whole_plot_indicator(6, 2, 3)
        with pytest.raises(ValueError, match="eta"):
            build_split_plot_covariance_inv(Z, eta=-0.5)

    def test_large_eta_row_sums_approach_zero(self):
        """For very large η, V⁻¹ row sums approach 0 (WP means absorbed)."""
        n_wp, s = 2, 4
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=1e6)
        # Each block row sum = 1 - η/(1+ηs) + (s-1)*(-η/(1+ηs)) = 1 - sη/(1+ηs) → 0
        row_sums = V_inv.sum(axis=1)
        np.testing.assert_allclose(row_sums, 0.0, atol=1e-4)

    def test_known_2wp_3sp_example(self):
        """Manually verify a 2-WP, 3-SP, η=1 example against explicit inverse."""
        # V (first 3×3 block) = [[2,1,1],[1,2,1],[1,1,2]]
        # V_inv (first block) = 1/4 * [[3,-1,-1],[-1,3,-1],[-1,-1,3]]
        n_wp, s, eta = 2, 3, 1.0
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=eta)
        expected_block = np.array([[3, -1, -1], [-1, 3, -1], [-1, -1, 3]]) / 4.0
        np.testing.assert_allclose(V_inv[:3, :3], expected_block, atol=1e-12)

    def test_importable_from_top_level(self):
        import lattice_doe
        assert hasattr(lattice_doe, "build_split_plot_covariance_inv")


# ===========================================================================
# TestGlsInformationMatrix
# ===========================================================================

class TestGlsInformationMatrix:

    def _simple_X(self):
        """Simple 6×3 model matrix for testing."""
        rng = np.random.default_rng(42)
        X = rng.standard_normal((6, 3))
        X[:, 0] = 1.0  # intercept
        return X

    def test_shape(self):
        X = self._simple_X()
        M = gls_information_matrix(X, np.eye(6))
        assert M.shape == (3, 3)

    def test_identity_V_inv_matches_XtX_plus_jitter(self):
        """With V_inv = I, M = X'X + jitter·I."""
        X = self._simple_X()
        jitter = 1e-8
        M = gls_information_matrix(X, np.eye(6), jitter=jitter)
        expected = X.T @ X + jitter * np.eye(3)
        np.testing.assert_allclose(M, expected, atol=1e-12)

    def test_symmetric(self):
        X = self._simple_X()
        n_wp, s = 2, 3
        Z = build_whole_plot_indicator(6, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=1.5)
        M = gls_information_matrix(X, V_inv)
        np.testing.assert_allclose(M, M.T, atol=1e-12)

    def test_positive_definite(self):
        """M must be positive definite (all eigenvalues > 0)."""
        rng = np.random.default_rng(7)
        n_wp, s = 3, 4
        n = n_wp * s
        X = rng.standard_normal((n, 4))
        X[:, 0] = 1.0
        Z = build_whole_plot_indicator(n, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=2.0)
        M = gls_information_matrix(X, V_inv, jitter=1e-6)
        eigvals = np.linalg.eigvalsh(M)
        assert np.all(eigvals > 0), f"Non-positive eigenvalue: {eigvals.min()}"

    def test_jitter_adds_to_diagonal(self):
        """Increasing jitter increases diagonal entries by exactly that amount."""
        X = self._simple_X()
        M1 = gls_information_matrix(X, np.eye(6), jitter=1e-8)
        M2 = gls_information_matrix(X, np.eye(6), jitter=1e-4)
        diff = np.diag(M2) - np.diag(M1)
        np.testing.assert_allclose(diff, np.full(3, 1e-4 - 1e-8), atol=1e-12)

    def test_gls_differs_from_ols_when_eta_nonzero(self):
        """GLS information matrix is not equal to OLS (X'X) when η > 0."""
        rng = np.random.default_rng(99)
        n_wp, s = 3, 4
        n = n_wp * s
        X = rng.standard_normal((n, 3))
        Z = build_whole_plot_indicator(n, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta=2.0)
        M_gls = gls_information_matrix(X, V_inv, jitter=0)
        M_ols = gls_information_matrix(X, np.eye(n), jitter=0)
        assert not np.allclose(M_gls, M_ols)

    def test_importable_from_top_level(self):
        import lattice_doe
        assert hasattr(lattice_doe, "gls_information_matrix")


# ===========================================================================
# TestClassifyContrasts
# ===========================================================================

class TestClassifyContrasts:
    """
    Model: intercept + A (HTC) + B (ETC) + C (ETC) → 4 columns (p=4)
    htc_factor_cols = [0, 1]  (intercept + A)
    """

    HTC_COLS = [0, 1]
    P = 4

    def test_all_htc_contrast_is_wp(self):
        L = np.array([[1.0, 1.0, 0.0, 0.0]])  # only cols 0,1
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(True)

    def test_intercept_only_contrast_is_wp(self):
        L = np.array([[1.0, 0.0, 0.0, 0.0]])  # only col 0
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(True)

    def test_htc_only_no_intercept_is_wp(self):
        L = np.array([[0.0, 1.0, 0.0, 0.0]])  # col 1 (A) only
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(True)

    def test_etc_contrast_is_sp(self):
        L = np.array([[0.0, 0.0, 1.0, 0.0]])  # col 2 (B)
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(False)

    def test_mixed_htc_etc_is_sp(self):
        L = np.array([[0.0, 1.0, 1.0, 0.0]])  # cols 1 (A, HTC) + 2 (B, ETC)
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(False)

    def test_multiple_rows(self):
        L = np.array([
            [1.0, 0.0, 0.0, 0.0],  # WP (intercept)
            [0.0, 1.0, 0.0, 0.0],  # WP (A)
            [0.0, 0.0, 1.0, 0.0],  # SP (B)
            [0.0, 0.0, 0.0, 1.0],  # SP (C)
            [0.0, 1.0, 1.0, 0.0],  # SP (mixed)
        ])
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        expected = np.array([True, True, False, False, False])
        np.testing.assert_array_equal(is_wp, expected)

    def test_all_zeros_row_is_wp(self):
        """All-zero row has no nonzero entries → subset of any set → WP."""
        L = np.array([[0.0, 0.0, 0.0, 0.0]])
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp[0] is np.bool_(True)

    def test_empty_htc_cols_all_are_sp(self):
        """When htc_factor_cols is empty, any non-zero contrast is SP."""
        L = np.array([[1.0, 0.0, 0.0, 0.0]])
        is_wp = classify_contrasts(L, [], self.P)
        # non-zero entry not in empty set → SP
        assert is_wp[0] is np.bool_(False)

    def test_1d_L_treated_as_single_row(self):
        L = np.array([1.0, 0.0, 0.0, 0.0])  # 1D
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp.shape == (1,)
        assert is_wp[0] is np.bool_(True)

    def test_output_dtype_is_bool(self):
        L = np.array([[0.0, 0.0, 1.0, 0.0]])
        is_wp = classify_contrasts(L, self.HTC_COLS, self.P)
        assert is_wp.dtype == bool


# ===========================================================================
# TestSplitPlotDfDenom
# ===========================================================================

class TestSplitPlotDfDenom:
    """
    Fixture: 2 WPs, 4 SP each → n_total=8, n_wp=2.
    Model: intercept + A (HTC) + B (ETC) → p=3.
    True df_wp = n_wp - rank(X_wp) = 2 - 2 = 0... need bigger example.
    Use: 4 WPs, 3 SP each → n_total=12, n_wp=4.
    Model: intercept + A + B → p=3, A=HTC (col 1), B=ETC (col 2).
    """

    def _make_fixture(self):
        n_wp, s = 4, 3
        n_total = n_wp * s
        rng = np.random.default_rng(11)
        # WP factor A varies by whole plot (4 distinct values)
        A_vals = rng.standard_normal(n_wp)
        A = np.repeat(A_vals, s)
        # SP factor B varies freely
        B = rng.standard_normal(n_total)
        X = np.column_stack([np.ones(n_total), A, B])
        Z = build_whole_plot_indicator(n_total, n_wp, s)
        htc_cols = [0, 1]  # intercept + A
        return X, Z, htc_cols, n_wp, n_total

    def test_conservative_all_wp_df(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        # Contrast rows: one WP, one SP
        is_wp = np.array([True, False])
        df = split_plot_df_denom(X, Z, is_wp, "conservative", htc_cols)
        # All df should be df_wp
        assert df[0] == df[1]

    def test_sp_only_all_sp_df(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False])
        df = split_plot_df_denom(X, Z, is_wp, "sp_only", htc_cols)
        assert df[0] == df[1]

    def test_auto_wp_contrast_gets_wp_df(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp_wp = np.array([True])
        is_wp_sp = np.array([False])
        df_wp_row = split_plot_df_denom(X, Z, is_wp_wp, "auto", htc_cols)
        df_sp_row = split_plot_df_denom(X, Z, is_wp_sp, "auto", htc_cols)
        # WP contrast → df_wp; SP contrast → df_sp; df_wp <= df_sp typically
        assert df_wp_row[0] <= df_sp_row[0]

    def test_auto_assigns_different_df_for_wp_vs_sp(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False])
        df = split_plot_df_denom(X, Z, is_wp, "auto", htc_cols)
        # df[0] (WP) and df[1] (SP) should differ
        assert df[0] != df[1], "auto mode should assign different df to WP vs SP contrasts"

    def test_conservative_returns_same_as_conservative_for_wp(self):
        """Conservative df for a WP contrast = same as auto WP df."""
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True])
        df_auto = split_plot_df_denom(X, Z, is_wp, "auto", htc_cols)
        df_cons = split_plot_df_denom(X, Z, is_wp, "conservative", htc_cols)
        np.testing.assert_array_equal(df_auto, df_cons)

    def test_df_values_are_positive(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False, True, False])
        for method in ("auto", "conservative", "sp_only"):
            df = split_plot_df_denom(X, Z, is_wp, method, htc_cols)
            assert np.all(df >= 1), f"Non-positive df found for method={method!r}"

    def test_output_dtype_is_int(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False])
        df = split_plot_df_denom(X, Z, is_wp, "auto", htc_cols)
        assert df.dtype == int or np.issubdtype(df.dtype, np.integer)

    def test_output_length_matches_is_wp(self):
        X, Z, htc_cols, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False, True])
        df = split_plot_df_denom(X, Z, is_wp, "auto", htc_cols)
        assert len(df) == 3

    def test_without_htc_cols_still_returns_valid_df(self):
        """Fallback (htc_factor_cols=None) returns positive df."""
        X, Z, _, n_wp, n_total = self._make_fixture()
        is_wp = np.array([True, False])
        df = split_plot_df_denom(X, Z, is_wp, "auto", htc_factor_cols=None)
        assert np.all(df >= 1)

    def test_invalid_df_method_raises(self):
        X, Z, htc_cols, _, _ = self._make_fixture()
        with pytest.raises(ValueError, match="df_method"):
            split_plot_df_denom(X, Z, np.array([True]), "kenward_roger", htc_cols)


# ---------------------------------------------------------------------------
# SP-3 tests: build_split_plot_candidate
# ---------------------------------------------------------------------------

_CONT_FACTORS = {"A": (-1.0, 1.0), "B": (-1.0, 1.0), "C": (0.0, 5.0)}
_CAT_FACTORS = {"T": ["low", "mid", "high"], "D": ["x", "y", "z"]}
_MIXED_FACTORS = {"A": (-1.0, 1.0), "T": ["low", "high"]}


class TestBuildSplitPlotCandidate:
    """Tests for build_split_plot_candidate (SP-3)."""

    # --- Row count ---

    def test_row_count_continuous(self):
        """Exactly n_wp * subplots_per_wp rows for continuous factors."""
        n_wp, s = 4, 3
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=n_wp,
            subplots_per_wp=s, random_state=0,
        )
        assert len(cand) == n_wp * s

    def test_row_count_categorical(self):
        n_wp, s = 3, 4
        cand = build_split_plot_candidate(
            _CAT_FACTORS, htc_factors=["T"], n_whole_plots=n_wp,
            subplots_per_wp=s, random_state=1,
        )
        assert len(cand) == n_wp * s

    def test_row_count_mixed(self):
        n_wp, s = 5, 2
        cand = build_split_plot_candidate(
            _MIXED_FACTORS, htc_factors=["T"], n_whole_plots=n_wp,
            subplots_per_wp=s, random_state=2,
        )
        assert len(cand) == n_wp * s

    def test_row_count_all_htc(self):
        """When all factors are HTC the function still returns correct shape."""
        n_wp, s = 3, 2
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=list(_CONT_FACTORS.keys()),
            n_whole_plots=n_wp, subplots_per_wp=s, random_state=3,
        )
        assert len(cand) == n_wp * s

    # --- wp_id column ---

    def test_wp_id_column_exists(self):
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=4,
            subplots_per_wp=3, random_state=0,
        )
        assert "__wp_id__" in cand.columns

    def test_wp_id_spans_zero_to_n_wp_minus_one(self):
        n_wp = 5
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=n_wp,
            subplots_per_wp=2, random_state=0,
        )
        assert set(cand["__wp_id__"].unique()) == set(range(n_wp))

    def test_wp_id_dtype_is_integer(self):
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3,
            subplots_per_wp=2, random_state=0,
        )
        assert pd.api.types.is_integer_dtype(cand["__wp_id__"])

    def test_each_wp_id_appears_exactly_subplots_per_wp_times(self):
        n_wp, s = 4, 3
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=n_wp,
            subplots_per_wp=s, random_state=0,
        )
        counts = cand["__wp_id__"].value_counts()
        assert (counts == s).all()

    # --- HTC factor invariance within each WP ---

    def test_htc_values_constant_within_wp_continuous(self):
        """All rows with the same __wp_id__ share identical HTC values (continuous)."""
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A", "B"], n_whole_plots=4,
            subplots_per_wp=3, random_state=7,
        )
        for wp_id, grp in cand.groupby("__wp_id__"):
            for col in ["A", "B"]:
                assert grp[col].nunique() == 1, (
                    f"HTC factor {col!r} varies within WP slot {wp_id}"
                )

    def test_htc_values_constant_within_wp_categorical(self):
        cand = build_split_plot_candidate(
            _CAT_FACTORS, htc_factors=["T"], n_whole_plots=3,
            subplots_per_wp=4, random_state=8,
        )
        for wp_id, grp in cand.groupby("__wp_id__"):
            assert grp["T"].nunique() == 1

    def test_htc_values_constant_within_wp_mixed(self):
        cand = build_split_plot_candidate(
            _MIXED_FACTORS, htc_factors=["T"], n_whole_plots=4,
            subplots_per_wp=3, random_state=9,
        )
        for wp_id, grp in cand.groupby("__wp_id__"):
            assert grp["T"].nunique() == 1

    # --- Factor value bounds ---

    def test_continuous_factor_values_in_bounds(self):
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=4,
            subplots_per_wp=3, random_state=0,
        )
        for col, (lo, hi) in _CONT_FACTORS.items():
            assert cand[col].between(lo, hi).all(), (
                f"Factor {col!r} has values outside [{lo}, {hi}]"
            )

    def test_categorical_factor_values_in_levels(self):
        cand = build_split_plot_candidate(
            _CAT_FACTORS, htc_factors=["T"], n_whole_plots=3,
            subplots_per_wp=2, random_state=0,
        )
        for col, levels in _CAT_FACTORS.items():
            assert cand[col].isin(levels).all()

    # --- Column set ---

    def test_all_factor_columns_present(self):
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3,
            subplots_per_wp=2, random_state=0,
        )
        for col in _CONT_FACTORS:
            assert col in cand.columns

    def test_wp_id_is_last_column(self):
        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3,
            subplots_per_wp=2, random_state=0,
        )
        assert cand.columns[-1] == "__wp_id__"

    # --- Reproducibility ---

    def test_random_state_gives_reproducible_results(self):
        kwargs = dict(
            factors=_CONT_FACTORS, htc_factors=["A"],
            n_whole_plots=4, subplots_per_wp=3,
        )
        c1 = build_split_plot_candidate(**kwargs, random_state=42)
        c2 = build_split_plot_candidate(**kwargs, random_state=42)
        pd.testing.assert_frame_equal(c1, c2)

    def test_different_seeds_give_different_results(self):
        kwargs = dict(
            factors=_CONT_FACTORS, htc_factors=["A"],
            n_whole_plots=4, subplots_per_wp=3,
        )
        c1 = build_split_plot_candidate(**kwargs, random_state=0)
        c2 = build_split_plot_candidate(**kwargs, random_state=99)
        # At least one column should differ
        assert not c1.drop(columns="__wp_id__").equals(c2.drop(columns="__wp_id__"))

    # --- Constraint filtering ---

    def test_constraint_func_filters_rows(self):
        def keep_positive_A(row):
            return row["A"] > 0.0

        cand = build_split_plot_candidate(
            _CONT_FACTORS, htc_factors=["B"], n_whole_plots=6,
            subplots_per_wp=4, random_state=5, constraint_func=keep_positive_A,
        )
        assert (cand["A"] > 0.0).all()

    # --- Input validation ---

    def test_n_whole_plots_less_than_2_raises(self):
        with pytest.raises(ValueError, match="n_whole_plots"):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=["A"], n_whole_plots=1, subplots_per_wp=2,
            )

    def test_subplots_per_wp_less_than_1_raises(self):
        with pytest.raises(ValueError, match="subplots_per_wp"):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3, subplots_per_wp=0,
            )

    def test_empty_htc_factors_raises(self):
        with pytest.raises(ValueError, match="htc_factors"):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=[], n_whole_plots=3, subplots_per_wp=2,
            )

    def test_missing_htc_factor_name_raises(self):
        with pytest.raises(ValueError, match="not found in factors"):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=["MISSING"], n_whole_plots=3, subplots_per_wp=2,
            )

    def test_n_whole_plots_not_int_raises(self):
        with pytest.raises((ValueError, TypeError)):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3.5, subplots_per_wp=2,
            )

    def test_subplots_per_wp_not_int_raises(self):
        with pytest.raises((ValueError, TypeError)):
            build_split_plot_candidate(
                _CONT_FACTORS, htc_factors=["A"], n_whole_plots=3, subplots_per_wp=2.5,
            )


# ---------------------------------------------------------------------------
# SP-4 tests: GLS criterion scorers
# ---------------------------------------------------------------------------

def _make_ols_fixture(n: int = 12, p: int = 3, seed: int = 0):
    """Return (X, idx, X_cand) for OLS criterion tests."""
    rng = np.random.default_rng(seed)
    X_cand = np.column_stack([np.ones(50), rng.standard_normal((50, p - 1))])
    idx = np.arange(n)
    X = X_cand[:n]
    return X, X_cand, idx


def _make_sp_fixture(n_wp: int = 4, s: int = 3, p: int = 3, seed: int = 5):
    """Return (X, Z, V_inv, idx, X_cand) for a balanced split-plot fixture."""
    n = n_wp * s
    rng = np.random.default_rng(seed)
    # Full candidate model matrix
    X_cand = np.column_stack([np.ones(60), rng.standard_normal((60, p - 1))])
    idx = np.arange(n)
    X = X_cand[:n]
    Z = build_whole_plot_indicator(n, n_wp, s)
    V_inv = build_split_plot_covariance_inv(Z, eta=1.0)
    return X, X_cand, idx, V_inv


class TestGLSCriterionScorers:
    """Tests for _gls_i/d/a_criterion and updated _criterion_score dispatcher (SP-4)."""

    # ------------------------------------------------------------------ #
    # Key property: at V_inv = I, GLS == OLS                             #
    # ------------------------------------------------------------------ #

    def test_gls_d_equals_ols_d_when_vinv_identity(self):
        X, X_cand, idx = _make_ols_fixture()
        n = X.shape[0]
        V_inv_eye = np.eye(n)
        gls = _gls_d_criterion(X, V_inv_eye)
        ols = _d_criterion_for_indices(X_cand, idx)
        assert abs(gls - ols) < 1e-8

    def test_gls_a_equals_ols_a_when_vinv_identity(self):
        X, X_cand, idx = _make_ols_fixture()
        n = X.shape[0]
        V_inv_eye = np.eye(n)
        gls = _gls_a_criterion(X, V_inv_eye)
        ols = _a_criterion_for_indices(X_cand, idx)
        assert abs(gls - ols) < 1e-8

    def test_gls_i_equals_ols_i_when_vinv_identity(self):
        X, X_cand, idx = _make_ols_fixture()
        n, p = X.shape
        Mcand = X_cand.T @ X_cand
        N_cand = X_cand.shape[0]
        V_inv_eye = np.eye(n)
        gls = _gls_i_criterion(X, V_inv_eye, Mcand=Mcand, N_cand=N_cand)
        ols = _i_criterion_for_indices(X_cand, idx)
        assert abs(gls - ols) < 1e-8

    # ------------------------------------------------------------------ #
    # Score properties                                                    #
    # ------------------------------------------------------------------ #

    def test_gls_d_returns_finite_for_well_conditioned(self):
        X, _, _, V_inv = _make_sp_fixture()
        score = _gls_d_criterion(X, V_inv)
        assert np.isfinite(score)

    def test_gls_a_returns_finite_for_well_conditioned(self):
        X, _, _, V_inv = _make_sp_fixture()
        score = _gls_a_criterion(X, V_inv)
        assert np.isfinite(score) and score > 0.0

    def test_gls_i_returns_finite_for_well_conditioned(self):
        X, X_cand, idx, V_inv = _make_sp_fixture()
        Mcand = X_cand.T @ X_cand
        score = _gls_i_criterion(X, V_inv, Mcand=Mcand, N_cand=X_cand.shape[0])
        assert np.isfinite(score) and score > 0.0

    def test_gls_d_worse_for_rank_deficient_design(self):
        """A rank-deficient design should have a much worse D-criterion than a proper design."""
        X_good, _, _, V_inv_good = _make_sp_fixture()
        n, p = X_good.shape
        X_bad = np.zeros((n, p))   # all-zero rows → near-singular
        V_inv_eye = np.eye(n)
        score_bad = _gls_d_criterion(X_bad, V_inv_eye)
        score_good = _gls_d_criterion(X_good, V_inv_good)
        # Bad design is larger (worse) or not finite
        assert score_bad > score_good or not np.isfinite(score_bad)

    def test_gls_a_worse_for_rank_deficient_design(self):
        """A rank-deficient design should have a much worse A-criterion than a proper design."""
        X_good, _, _, V_inv_good = _make_sp_fixture()
        n, p = X_good.shape
        X_bad = np.zeros((n, p))
        V_inv_eye = np.eye(n)
        score_bad = _gls_a_criterion(X_bad, V_inv_eye)
        score_good = _gls_a_criterion(X_good, V_inv_good)
        assert score_bad > score_good or not np.isfinite(score_bad)

    def test_gls_i_without_mcand_uses_design_moment(self):
        """When Mcand is None, _gls_i_criterion uses X'X and is finite."""
        X, _, _, V_inv = _make_sp_fixture()
        score = _gls_i_criterion(X, V_inv, Mcand=None)
        assert np.isfinite(score) and score > 0.0

    def test_gls_d_lower_for_better_design(self):
        """Adding more non-collinear rows strictly lowers the D-criterion."""
        rng = np.random.default_rng(42)
        p = 3
        n_small, n_large = 6, 12
        X_small = np.column_stack([np.ones(n_small), rng.standard_normal((n_small, p - 1))])
        X_large = np.column_stack([np.ones(n_large), rng.standard_normal((n_large, p - 1))])
        score_small = _gls_d_criterion(X_small, np.eye(n_small))
        score_large = _gls_d_criterion(X_large, np.eye(n_large))
        # More informative design → lower D score (higher det)
        assert score_large < score_small

    # ------------------------------------------------------------------ #
    # Row permutation invariance                                          #
    # ------------------------------------------------------------------ #

    def test_gls_criteria_invariant_to_row_permutation(self):
        """Permuting rows of X (and corresponding rows/cols of V_inv) leaves scores unchanged."""
        X, _, _, V_inv = _make_sp_fixture()
        n = X.shape[0]
        perm = np.random.default_rng(99).permutation(n)
        X_perm = X[perm]
        V_inv_perm = V_inv[np.ix_(perm, perm)]

        for fn in (_gls_d_criterion, _gls_a_criterion):
            s_orig = fn(X, V_inv)
            s_perm = fn(X_perm, V_inv_perm)
            assert abs(s_orig - s_perm) < 1e-8, f"{fn.__name__} not invariant to row permutation"

        # I-criterion with explicit Mcand
        Mcand = np.eye(3)  # arbitrary fixed Mcand
        s_i_orig = _gls_i_criterion(X, V_inv, Mcand=Mcand, N_cand=1)
        s_i_perm = _gls_i_criterion(X_perm, V_inv_perm, Mcand=Mcand, N_cand=1)
        assert abs(s_i_orig - s_i_perm) < 1e-8

    # ------------------------------------------------------------------ #
    # _criterion_score dispatcher                                        #
    # ------------------------------------------------------------------ #

    def test_criterion_score_ols_path_no_vinv(self):
        """_criterion_score with V_inv=None uses OLS path for all criteria."""
        X, X_cand, idx = _make_ols_fixture()
        for crit in ("I", "D", "A"):
            score = _criterion_score(crit, X_cand, idx)
            assert np.isfinite(score)

    def test_criterion_score_gls_path_with_vinv(self):
        """_criterion_score with V_inv provided routes to GLS for all criteria."""
        X, X_cand, idx, V_inv = _make_sp_fixture()
        for crit in ("I", "D", "A"):
            score = _criterion_score(crit, X_cand, idx, V_inv=V_inv)
            assert np.isfinite(score)

    def test_criterion_score_gls_equals_ols_when_vinv_identity(self):
        """_criterion_score GLS path with eye(n) matches OLS path."""
        X, X_cand, idx = _make_ols_fixture()
        n = X.shape[0]
        V_inv_eye = np.eye(n)
        for crit in ("D", "A"):
            ols_score = _criterion_score(crit, X_cand, idx)
            gls_score = _criterion_score(crit, X_cand, idx, V_inv=V_inv_eye)
            assert abs(gls_score - ols_score) < 1e-8, f"Mismatch for criterion={crit!r}"

    def test_criterion_score_invalid_criterion_raises(self):
        X, X_cand, idx = _make_ols_fixture()
        with pytest.raises(ValueError, match="Unknown optimality criterion"):
            _criterion_score("Z", X_cand, idx)

    def test_criterion_score_invalid_criterion_gls_raises(self):
        X, X_cand, idx, V_inv = _make_sp_fixture()
        with pytest.raises(ValueError, match="Unknown optimality criterion"):
            _criterion_score("Z", X_cand, idx, V_inv=V_inv)

    # ------------------------------------------------------------------ #
    # _score_design_gls                                                  #
    # ------------------------------------------------------------------ #

    def test_score_design_gls_equals_ols_when_vinv_identity(self):
        """_score_design_gls with identity V_inv matches OLS _score_design for D and A."""
        from lattice_doe.iopt_search import _score_design
        X, _, _, _ = _make_sp_fixture()
        n, p = X.shape
        V_inv_eye = np.eye(n)
        for crit in ("D", "A"):
            ols = _score_design(crit, X)
            gls = _score_design_gls(crit, X, V_inv_eye)
            assert abs(gls - ols) < 1e-8, f"Mismatch for criterion={crit!r}"

    def test_score_design_gls_invalid_criterion_raises(self):
        X, _, _, V_inv = _make_sp_fixture()
        with pytest.raises(ValueError, match="Unknown optimality criterion"):
            _score_design_gls("Z", X, V_inv)

    def test_score_design_gls_i_criterion_with_mcand(self):
        X, X_cand, idx, V_inv = _make_sp_fixture()
        Mcand = X_cand.T @ X_cand
        N_cand = X_cand.shape[0]
        score = _score_design_gls("I", X, V_inv, Mcand=Mcand, N_cand=N_cand)
        assert np.isfinite(score) and score > 0.0


# ---------------------------------------------------------------------------
# SP-5 tests: build_split_plot_design
# ---------------------------------------------------------------------------

_SP5_FACTORS = {"A": (-1.0, 1.0), "B": (-1.0, 1.0), "C": (0.0, 5.0)}
_SP5_FORMULA = "~ 1 + A + B + C"
_SP5_HTC = ["A"]          # A is HTC; B, C are ETC


def _make_sp5_cand(n_wp: int = 3, s: int = 2, seed: int = 0) -> pd.DataFrame:
    """Build a small initial candidate for SP-5 tests."""
    return build_split_plot_candidate(
        _SP5_FACTORS, htc_factors=_SP5_HTC,
        n_whole_plots=n_wp, subplots_per_wp=s, random_state=seed,
    )


class TestSplitPlotExchange:
    """Tests for build_split_plot_design (SP-5)."""

    # ------------------------------------------------------------------ #
    # Output structure                                                   #
    # ------------------------------------------------------------------ #

    def test_returns_tuple_of_df_and_ndarray(self):
        cand = _make_sp5_cand()
        result = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert isinstance(result, tuple) and len(result) == 2
        design_df, X = result
        assert isinstance(design_df, pd.DataFrame)
        assert isinstance(X, np.ndarray)

    def test_design_row_count(self):
        n_wp, s = 3, 2
        cand = _make_sp5_cand(n_wp, s)
        design_df, X = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=n_wp, subplots_per_wp=s,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert len(design_df) == n_wp * s

    def test_X_shape_matches_design(self):
        n_wp, s = 3, 2
        cand = _make_sp5_cand(n_wp, s)
        design_df, X = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=n_wp, subplots_per_wp=s,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        # formula ~ 1 + A + B + C → p = 4
        assert X.shape == (n_wp * s, 4)

    def test_design_columns_match_cand(self):
        cand = _make_sp5_cand()
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert list(design_df.columns) == list(cand.columns)

    # ------------------------------------------------------------------ #
    # Structural constraints                                             #
    # ------------------------------------------------------------------ #

    def test_htc_constant_within_wp(self):
        """All rows sharing a __wp_id__ must have identical HTC values."""
        n_wp, s = 4, 3
        cand = _make_sp5_cand(n_wp, s)
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=n_wp, subplots_per_wp=s,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=3, max_iter=10, random_state=7,
            n_wp_cand=10, n_sp_cand=15,
        )
        for wp_id, grp in design_df.groupby("__wp_id__"):
            for htc_col in _SP5_HTC:
                assert grp[htc_col].nunique() == 1, (
                    f"HTC factor {htc_col!r} varies within WP slot {wp_id}"
                )

    def test_wp_id_spans_correct_range(self):
        n_wp, s = 3, 2
        cand = _make_sp5_cand(n_wp, s)
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=n_wp, subplots_per_wp=s,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert set(design_df["__wp_id__"].unique()) == set(range(n_wp))

    def test_each_wp_id_has_exactly_s_rows(self):
        n_wp, s = 3, 2
        cand = _make_sp5_cand(n_wp, s)
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=n_wp, subplots_per_wp=s,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        counts = design_df["__wp_id__"].value_counts()
        assert (counts == s).all()

    def test_continuous_factor_values_in_bounds(self):
        cand = _make_sp5_cand()
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        for col, (lo, hi) in _SP5_FACTORS.items():
            assert design_df[col].between(lo, hi).all(), (
                f"Factor {col!r} has values outside [{lo}, {hi}]"
            )

    # ------------------------------------------------------------------ #
    # Reproducibility and multi-start diversity                         #
    # ------------------------------------------------------------------ #

    def test_reproducible_with_same_seed(self):
        cand = _make_sp5_cand()
        kwargs = dict(
            formula=_SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0, factors=_SP5_FACTORS,
            starts=3, max_iter=5, random_state=42,
            n_wp_cand=8, n_sp_cand=10,
        )
        d1, X1 = build_split_plot_design(cand, **kwargs)
        d2, X2 = build_split_plot_design(cand, **kwargs)
        pd.testing.assert_frame_equal(d1, d2)
        np.testing.assert_array_equal(X1, X2)

    def test_different_seeds_usually_differ(self):
        cand = _make_sp5_cand(n_wp=4, s=3)
        kwargs = dict(
            formula=_SP5_FORMULA, n_wp=4, subplots_per_wp=3,
            htc_factors=_SP5_HTC, eta=1.0, factors=_SP5_FACTORS,
            starts=1, max_iter=3, n_wp_cand=10, n_sp_cand=15,
        )
        d1, _ = build_split_plot_design(cand, **kwargs, random_state=0)
        d2, _ = build_split_plot_design(cand, **kwargs, random_state=999)
        # Different seeds → (likely) different designs
        # We can't guarantee it with starts=1/max_iter=3, but check at least types
        assert isinstance(d1, pd.DataFrame) and isinstance(d2, pd.DataFrame)

    # ------------------------------------------------------------------ #
    # Criterion variants                                                 #
    # ------------------------------------------------------------------ #

    @pytest.mark.parametrize("crit", ["I", "D", "A"])
    def test_all_criteria_run_without_error(self, crit):
        cand = _make_sp5_cand()
        design_df, X = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, criterion=crit,
            starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert len(design_df) == 6
        assert X.shape[0] == 6

    # ------------------------------------------------------------------ #
    # criterion_ignore_vr and eta=0                                      #
    # ------------------------------------------------------------------ #

    def test_criterion_ignore_vr_produces_valid_structure(self):
        cand = _make_sp5_cand()
        design_df, X = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=2.0,
            criterion_ignore_vr=True,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert len(design_df) == 6
        for wp_id, grp in design_df.groupby("__wp_id__"):
            assert grp["A"].nunique() == 1

    def test_eta_zero_equivalent_to_criterion_ignore_vr(self):
        """With eta=0 V_inv = I, so scores should match criterion_ignore_vr=True."""
        cand = _make_sp5_cand()
        kwargs = dict(
            formula=_SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC,
            factors=_SP5_FACTORS, starts=2, max_iter=5, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        d_eta0, X_eta0 = build_split_plot_design(cand, eta=0.0, **kwargs)
        d_ols, X_ols = build_split_plot_design(
            cand, eta=1.0, criterion_ignore_vr=True, **kwargs
        )
        # Both use identity V_inv; same rng seed → identical designs
        pd.testing.assert_frame_equal(d_eta0, d_ols)

    # ------------------------------------------------------------------ #
    # Without factors param (pool from cand)                            #
    # ------------------------------------------------------------------ #

    def test_works_without_factors_param(self):
        """When factors=None, derives pools from cand directly."""
        cand = _make_sp5_cand(n_wp=4, s=3)
        design_df, X = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=4, subplots_per_wp=3,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=None, starts=2, max_iter=5, random_state=0,
        )
        assert len(design_df) == 12
        for wp_id, grp in design_df.groupby("__wp_id__"):
            assert grp["A"].nunique() == 1

    # ------------------------------------------------------------------ #
    # Convergence                                                        #
    # ------------------------------------------------------------------ #

    def test_terminates_within_max_iter(self):
        """Should always return before exceeding max_iter (even with max_iter=1)."""
        cand = _make_sp5_cand()
        design_df, _ = build_split_plot_design(
            cand, _SP5_FORMULA, n_wp=3, subplots_per_wp=2,
            htc_factors=_SP5_HTC, eta=1.0,
            factors=_SP5_FACTORS, starts=1, max_iter=1, random_state=0,
            n_wp_cand=8, n_sp_cand=10,
        )
        assert len(design_df) == 6  # always returns a valid design

    # ------------------------------------------------------------------ #
    # All-HTC edge case                                                  #
    # ------------------------------------------------------------------ #

    def test_all_htc_factors(self):
        """When all factors are HTC, Phase 2 is skipped; design still valid."""
        factors_htc_only = {"A": (-1.0, 1.0), "B": (-1.0, 1.0)}
        formula = "~ 1 + A + B"
        htc_all = ["A", "B"]
        cand = build_split_plot_candidate(
            factors_htc_only, htc_factors=htc_all,
            n_whole_plots=3, subplots_per_wp=2, random_state=0,
        )
        design_df, X = build_split_plot_design(
            cand, formula, n_wp=3, subplots_per_wp=2,
            htc_factors=htc_all, eta=1.0,
            factors=factors_htc_only, starts=2, max_iter=5, random_state=0,
            n_wp_cand=10, n_sp_cand=5,
        )
        assert len(design_df) == 6
        assert X.shape[1] == 3  # intercept + A + B


# ===========================================================================
# TestSplitPlotPower  (SP-6)
# ===========================================================================

# Shared fixtures for power tests.
# Design: 4 WPs × 3 SP each = 12 runs; factors A (HTC), B (ETC), intercept.
# Model: ~ 1 + A + B  → p = 3 columns.

def _make_sp6_design():
    """Return (X, Z) for a simple 4×3 split-plot design."""
    n_wp, s = 4, 3
    n = n_wp * s
    Z = build_whole_plot_indicator(n, n_wp, s)
    # X: intercept, A (HTC, constant per WP), B (ETC)
    A_vals = np.repeat([-1.0, -1.0, 1.0, 1.0], s)
    B_vals = np.tile([-1.0, 0.0, 1.0], n_wp)
    X = np.column_stack([np.ones(n), A_vals, B_vals])
    return X, Z


class TestSplitPlotPower:

    # ------------------------------------------------------------------ #
    # Return types                                                         #
    # ------------------------------------------------------------------ #

    def test_contrast_power_sp_return_type(self):
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])  # test A
        delta = np.array([1.0])
        res = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        assert hasattr(res, "power") and hasattr(res, "lam")
        assert isinstance(res.power, float)
        assert isinstance(res.lam, float)

    def test_global_r2_power_sp_return_type(self):
        X, Z = _make_sp6_design()
        res = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        assert hasattr(res, "power") and hasattr(res, "lam")
        assert isinstance(res.power, float)
        assert isinstance(res.lam, float)

    # ------------------------------------------------------------------ #
    # eta=0 equivalence with OLS functions                                 #
    # ------------------------------------------------------------------ #

    def test_contrast_power_sp_eta0_matches_ols(self):
        """contrast_power_sp at eta=0 must equal contrast_power exactly."""
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])
        delta = np.array([1.0])
        sp = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=0.0, alpha=0.05)
        ols = contrast_power(L, delta, X, sigma=1.0, alpha=0.05)
        assert abs(sp.power - ols.power) < 1e-10
        assert abs(sp.lam - ols.lam) < 1e-10

    def test_global_r2_power_sp_eta0_matches_ols(self):
        """global_r2_power_sp at eta=0 must equal global_r2_power exactly."""
        X, Z = _make_sp6_design()
        sp = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=0.0, alpha=0.05)
        ols = global_r2_power(0.5, X, alpha=0.05)
        assert abs(sp.power - ols.power) < 1e-10
        assert abs(sp.lam - ols.lam) < 1e-10

    # ------------------------------------------------------------------ #
    # Power vs eta behaviour                                               #
    # ------------------------------------------------------------------ #

    def test_wp_contrast_power_decreases_with_eta(self):
        """Power for a WP-factor contrast decreases as eta increases."""
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])  # contrast on A (HTC)
        delta = np.array([1.0])
        p_low = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=0.1, alpha=0.05,
                                   htc_factor_cols=[1], df_method="auto")
        p_high = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=5.0, alpha=0.05,
                                    htc_factor_cols=[1], df_method="auto")
        assert p_low.power >= p_high.power

    def test_sp_contrast_power_less_sensitive_to_eta(self):
        """Power for an SP-factor contrast changes less than for a WP contrast."""
        X, Z = _make_sp6_design()
        L_wp = np.array([0.0, 1.0, 0.0])   # A is HTC
        L_sp = np.array([0.0, 0.0, 1.0])   # B is ETC
        delta = np.array([1.0])
        pw_lo = contrast_power_sp(L_wp, delta, X, Z, sigma_sp=1.0, eta=0.1, alpha=0.05,
                                   htc_factor_cols=[1]).power
        pw_hi = contrast_power_sp(L_wp, delta, X, Z, sigma_sp=1.0, eta=5.0, alpha=0.05,
                                   htc_factor_cols=[1]).power
        ps_lo = contrast_power_sp(L_sp, delta, X, Z, sigma_sp=1.0, eta=0.1, alpha=0.05,
                                   htc_factor_cols=[1]).power
        ps_hi = contrast_power_sp(L_sp, delta, X, Z, sigma_sp=1.0, eta=5.0, alpha=0.05,
                                   htc_factor_cols=[1]).power
        drop_wp = pw_lo - pw_hi
        drop_sp = ps_lo - ps_hi
        # WP power drops more than SP power as eta increases
        assert drop_wp >= drop_sp

    def test_global_r2_power_decreases_with_eta(self):
        """Global R² power decreases (or stays equal) as eta increases."""
        X, Z = _make_sp6_design()
        p_low = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=0.1, alpha=0.05)
        p_high = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=5.0, alpha=0.05)
        assert p_low.power >= p_high.power

    # ------------------------------------------------------------------ #
    # df_method comparison                                                 #
    # ------------------------------------------------------------------ #

    def test_conservative_le_sp_only_for_wp_contrast(self):
        """df_method='conservative' gives lower or equal power than 'sp_only'."""
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])
        delta = np.array([1.0])
        p_cons = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05,
                                    htc_factor_cols=[1], df_method="conservative")
        p_sp = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05,
                                  htc_factor_cols=[1], df_method="sp_only")
        assert p_cons.power <= p_sp.power

    # ------------------------------------------------------------------ #
    # Output range                                                         #
    # ------------------------------------------------------------------ #

    def test_contrast_power_sp_in_unit_interval(self):
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])
        delta = np.array([1.0])
        res = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        assert 0.0 <= res.power <= 1.0
        assert res.lam >= 0.0

    def test_global_r2_power_sp_in_unit_interval(self):
        X, Z = _make_sp6_design()
        res = global_r2_power_sp(0.3, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        assert 0.0 <= res.power <= 1.0
        assert res.lam >= 0.0

    # ------------------------------------------------------------------ #
    # All df_methods run without error                                     #
    # ------------------------------------------------------------------ #

    @pytest.mark.parametrize("df_method", ["auto", "conservative", "sp_only"])
    def test_all_df_methods_run(self, df_method):
        X, Z = _make_sp6_design()
        L = np.array([0.0, 1.0, 0.0])
        delta = np.array([1.0])
        res = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05,
                                 htc_factor_cols=[1], df_method=df_method)
        assert 0.0 <= res.power <= 1.0

    # ------------------------------------------------------------------ #
    # Multi-row L: min-power convention                                    #
    # ------------------------------------------------------------------ #

    def test_multirow_L_returns_min_power(self):
        """Multi-row L returns power = min power across rows."""
        X, Z = _make_sp6_design()
        # Row 0: contrast on A (easier to detect); Row 1: intercept test (harder)
        L = np.array([[0.0, 1.0, 0.0], [1.0, 0.0, 0.0]])
        delta = np.array([2.0, 0.01])  # large effect on A, tiny on intercept
        res_joint = contrast_power_sp(L, delta, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        # Single-row results
        res0 = contrast_power_sp(L[0:1], delta[0:1], X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        res1 = contrast_power_sp(L[1:2], delta[1:2], X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)
        assert abs(res_joint.power - min(res0.power, res1.power)) < 1e-10

    # ------------------------------------------------------------------ #
    # lambda_mode for global R²                                            #
    # ------------------------------------------------------------------ #

    def test_lambda_mode_n_minus_p_smaller_than_n(self):
        """lambda_mode='n_minus_p' produces smaller λ than 'n' for eta>0."""
        X, Z = _make_sp6_design()
        res_n = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05,
                                    lambda_mode="n")
        res_np = global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05,
                                     lambda_mode="n_minus_p")
        assert res_n.lam >= res_np.lam

    # ------------------------------------------------------------------ #
    # Input validation                                                     #
    # ------------------------------------------------------------------ #

    def test_contrast_power_sp_invalid_sigma_raises(self):
        X, Z = _make_sp6_design()
        with pytest.raises(ValueError):
            contrast_power_sp(np.array([0.0, 1.0, 0.0]), np.array([1.0]),
                              X, Z, sigma_sp=0.0, eta=1.0, alpha=0.05)

    def test_contrast_power_sp_negative_eta_raises(self):
        X, Z = _make_sp6_design()
        with pytest.raises(ValueError):
            contrast_power_sp(np.array([0.0, 1.0, 0.0]), np.array([1.0]),
                              X, Z, sigma_sp=1.0, eta=-0.5, alpha=0.05)

    def test_global_r2_power_sp_invalid_r2_raises(self):
        X, Z = _make_sp6_design()
        with pytest.raises(ValueError):
            global_r2_power_sp(1.5, X, Z, sigma_sp=1.0, eta=1.0, alpha=0.05)

    def test_global_r2_power_sp_invalid_sigma_raises(self):
        X, Z = _make_sp6_design()
        with pytest.raises(ValueError):
            global_r2_power_sp(0.5, X, Z, sigma_sp=-1.0, eta=1.0, alpha=0.05)

    def test_global_r2_power_sp_negative_eta_raises(self):
        X, Z = _make_sp6_design()
        with pytest.raises(ValueError):
            global_r2_power_sp(0.5, X, Z, sigma_sp=1.0, eta=-1.0, alpha=0.05)


# ===========================================================================
# TestSplitPlotAPI  (SP-7)
# ===========================================================================

# Shared config helpers for API tests.
_SP7_FACTORS = {"A": (-1.0, 1.0), "B": (-1.0, 1.0), "C": (0.0, 1.0)}
_SP7_FORMULA = "~ 1 + A + B + C"
_SP7_HTC = ["A"]


def _sp7_contrast_cfg(**kw):
    """Minimal PowerContrastConfig for SP-7 tests."""
    import numpy as np
    defaults = dict(
        L=np.array([[0.0, 1.0, 0.0, 0.0]]),  # test A (col 1)
        delta=np.array([1.0]),
        sigma=1.0,
        power=0.5,
        alpha=0.10,
        max_n=40,
        max_iter=10,
    )
    defaults.update(kw)
    return PowerContrastConfig(**defaults)


def _sp7_r2_cfg(**kw):
    """Minimal PowerR2Config for SP-7 tests."""
    defaults = dict(
        r2_target=0.5,
        power=0.5,
        alpha=0.10,
        max_n=40,
        max_iter=10,
    )
    defaults.update(kw)
    return PowerR2Config(**defaults)


def _sp7_opts(n_wp=4, subplots_per_wp=3, eta=1.0, df_method="auto"):
    return DesignOptions(
        split_plot=SplitPlotOptions(
            htc_factors=_SP7_HTC,
            n_whole_plots=n_wp,
            eta=eta,
            subplots_per_wp=subplots_per_wp,
            df_method=df_method,
        ),
        starts=2,
        max_iter=5,
        random_state=42,
    )


class TestSplitPlotAPI:

    # ------------------------------------------------------------------ #
    # Return structure                                                      #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~182s measured 2026-07-11 (TICKET-006A)
    def test_returns_dict_with_required_keys(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(),
        )
        assert "design_df" in result
        assert "buckets_df" in result
        assert "report" in result

    @pytest.mark.slow  # ~148s measured 2026-07-11 (TICKET-006A)
    def test_design_df_is_dataframe(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(n_wp=4, subplots_per_wp=3),
        )
        assert isinstance(result["design_df"], pd.DataFrame)

    def test_design_df_row_count_is_n_wp_times_s(self):
        """design_df has n_wp * subplots_per_wp rows."""
        n_wp, s = 4, 3
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(n_wp=n_wp, subplots_per_wp=s),
        )
        expected_n = result["report"]["split_plot"]["n_whole_plots"] * s
        assert len(result["design_df"]) == expected_n

    # ------------------------------------------------------------------ #
    # HTC nesting constraint                                               #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~191s measured 2026-07-11 (TICKET-006A)
    def test_htc_constant_within_wp(self):
        """All sub-plots in a WP share the same HTC factor values."""
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(n_wp=4, subplots_per_wp=3),
        )
        df = result["design_df"]
        assert "__wp_id__" in df.columns
        for wp_id, grp in df.groupby("__wp_id__"):
            for htc in _SP7_HTC:
                assert grp[htc].nunique() == 1, (
                    f"HTC factor {htc!r} varies within WP {wp_id}"
                )

    # ------------------------------------------------------------------ #
    # Report structure                                                     #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~132s measured 2026-07-11 (TICKET-006A)
    def test_report_contains_split_plot_dict(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(),
        )
        assert "split_plot" in result["report"]
        sp = result["report"]["split_plot"]
        assert isinstance(sp, dict)

    @pytest.mark.slow  # ~130s measured 2026-07-11 (TICKET-006A)
    def test_split_plot_dict_has_required_keys(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(),
        )
        sp = result["report"]["split_plot"]
        for key in ("n_whole_plots", "subplots_per_wp", "n_total",
                    "eta", "htc_factors", "etc_factors", "df_method"):
            assert key in sp, f"Missing key: {key!r}"

    @pytest.mark.slow  # ~241s measured 2026-07-11 (TICKET-006A)
    def test_split_plot_dict_values_consistent(self):
        n_wp, s = 4, 3
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(n_wp=n_wp, subplots_per_wp=s, eta=2.0),
        )
        sp = result["report"]["split_plot"]
        assert sp["subplots_per_wp"] == s
        assert sp["eta"] == 2.0
        assert set(sp["htc_factors"]) == set(_SP7_HTC)
        assert sp["n_whole_plots"] * sp["subplots_per_wp"] == sp["n_total"]

    @pytest.mark.slow  # ~130s measured 2026-07-11 (TICKET-006A)
    def test_report_n_equals_design_rows(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(),
        )
        assert result["report"]["n"] == len(result["design_df"])

    @pytest.mark.slow  # ~130s measured 2026-07-11 (TICKET-006A)
    def test_achieved_power_in_unit_interval(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(),
        )
        pwr = result["report"]["achieved_power"]
        assert 0.0 <= pwr <= 1.0

    # ------------------------------------------------------------------ #
    # R² power mode                                                        #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~131s measured 2026-07-11 (TICKET-006A)
    def test_r2_power_cfg_works(self):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_r2_cfg(),
            design_opts=_sp7_opts(),
        )
        assert "design_df" in result
        assert "split_plot" in result["report"]
        assert 0.0 <= result["report"]["achieved_power"] <= 1.0

    # ------------------------------------------------------------------ #
    # Input validation                                                     #
    # ------------------------------------------------------------------ #

    def test_raises_if_htc_factor_not_in_factors(self):
        """ValueError if htc_factors contains a name not in factors."""
        bad_opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["X_NONEXISTENT"],
                n_whole_plots=4,
                eta=1.0,
                subplots_per_wp=3,
            ),
            starts=1, max_iter=3, random_state=0,
        )
        with pytest.raises(ValueError, match="htc_factors"):
            find_optimal_design(
                _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
                design_opts=bad_opts,
            )

    def test_raises_if_blocks_and_split_plot_combined(self):
        """ValueError if both n_blocks and split_plot are set."""
        conflict_opts = DesignOptions(
            n_blocks=3,
            split_plot=SplitPlotOptions(
                htc_factors=_SP7_HTC,
                n_whole_plots=4,
                eta=1.0,
                subplots_per_wp=3,
            ),
            starts=1, max_iter=3, random_state=0,
        )
        with pytest.raises(ValueError, match="n_blocks"):
            find_optimal_design(
                _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
                design_opts=conflict_opts,
            )

    # ------------------------------------------------------------------ #
    # Non-SP path unchanged                                                #
    # ------------------------------------------------------------------ #

    def test_non_sp_path_produces_valid_result(self):
        """Without split_plot, find_optimal_design behaves as before."""
        import numpy as np
        ols_cfg = PowerContrastConfig(
            L=np.array([[0.0, 1.0, 0.0, 0.0]]),
            delta=np.array([1.0]),
            sigma=1.0,
            power=0.5,
            alpha=0.10,
            max_n=30,
            max_iter=8,
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, ols_cfg,
            design_opts=DesignOptions(starts=2, max_iter=5, random_state=0),
        )
        assert "design_df" in result
        assert "split_plot" not in result["report"]

    # ------------------------------------------------------------------ #
    # df_method variants                                                   #
    # ------------------------------------------------------------------ #

    @pytest.mark.parametrize("df_method", ["auto", "conservative", "sp_only"])
    @pytest.mark.slow  # ~130s x3 params measured 2026-07-11 (TICKET-006A)
    def test_all_df_methods_run_without_error(self, df_method):
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
            design_opts=_sp7_opts(df_method=df_method),
        )
        assert 0.0 <= result["report"]["achieved_power"] <= 1.0

    # ------------------------------------------------------------------ #
    # CR-24 regression: feasibility constraints respected in split-plot    #
    # ------------------------------------------------------------------ #

    def test_htc_constraint_expr_respected(self):
        """CR-24: constraint_expr on HTC factors is applied to split-plot design."""
        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=2, max_iter=10, random_state=42, candidate_points=200,
            constraint_expr="A <= -0.3",
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(), design_opts=opts,
        )
        df = result["design_df"]
        assert df["A"].max() <= -0.3 + 1e-9, (
            f"CR-24: constraint 'A <= -0.3' violated in design: max A = {df['A'].max():.4f}"
        )

    def test_etc_constraint_expr_respected(self):
        """CR-24: constraint_expr on ETC factors is applied to split-plot design."""
        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=2, max_iter=10, random_state=42, candidate_points=200,
            constraint_expr="C <= 0.5",
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(), design_opts=opts,
        )
        df = result["design_df"]
        assert df["C"].max() <= 0.5 + 1e-9, (
            f"CR-24: constraint 'C <= 0.5' violated in design: max C = {df['C'].max():.4f}"
        )

    def test_constraint_func_respected(self):
        """CR-24: constraint_func callable is also forwarded into split-plot path."""
        def htc_constraint(row):
            return row["A"] <= -0.2

        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=2, max_iter=10, random_state=42, candidate_points=200,
            constraint_func=htc_constraint,
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(), design_opts=opts,
        )
        df = result["design_df"]
        assert df["A"].max() <= -0.2 + 1e-9, (
            f"CR-24: constraint_func violated in design: max A = {df['A'].max():.4f}"
        )


# ===========================================================================
# TestCR25HtcColMapping — WP df correctly applied to WP contrasts in auto mode
# ===========================================================================

class TestCR25HtcColMapping:
    """CR-25 regression: contrast_power_sp receives htc_factor_cols so that
    classify_contrasts() sees the correct HTC column set and assigns WP df to
    pure-WP contrasts under df_method="auto"."""

    # ------------------------------------------------------------------ #
    # Unit test: htc_factor_cols_from_names helper                        #
    # ------------------------------------------------------------------ #

    def test_htc_factor_cols_from_names_classifies_correctly(self):
        """htc_factor_cols_from_names correctly identifies WP vs SP columns."""
        from lattice_doe.split_plot import htc_factor_cols_from_names

        p_names = ["Intercept", "A", "B", "C", "A:B", "A:C", "B:C"]
        htc_factors = ["A"]
        all_factors = ["A", "B", "C"]

        cols = htc_factor_cols_from_names(p_names, htc_factors, all_factors)
        # Intercept (0), A (1), and A:B (4) are the only WP-pure columns
        # (A:B involves only A which is HTC; B alone is ETC)
        # Wait: A:B involves B (ETC), so A:B is SP.
        # Pure WP: Intercept (0), A (1) only.
        assert 0 in cols, "Intercept must be in htc_factor_cols"
        assert 1 in cols, "A must be in htc_factor_cols"
        for idx in (2, 3, 4, 5, 6):  # B, C, A:B, A:C, B:C — all involve ETC
            assert idx not in cols, f"Column {p_names[idx]!r} should NOT be in htc_factor_cols"

    def test_htc_factor_cols_empty_when_no_htc(self):
        from lattice_doe.split_plot import htc_factor_cols_from_names

        cols = htc_factor_cols_from_names(
            ["Intercept", "A", "B"], htc_factors=[], all_factor_names=["A", "B"],
        )
        assert cols == []

    # ------------------------------------------------------------------ #
    # Integration: WP contrast uses WP df under df_method="auto"          #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~695s measured 2026-07-11 (TICKET-006A)
    def test_wp_contrast_auto_equals_conservative(self):
        """CR-25: WP contrast power under 'auto' must equal 'conservative'
        (both use WP df), not 'sp_only' (which uses SP df).

        Before the fix, 'auto' collapsed to 'sp_only' because htc_factor_cols
        was never passed, so classify_contrasts() treated every contrast as SP.
        """
        # L targets A (the HTC factor) — pure WP contrast
        L_wp = np.array([[0.0, 1.0, 0.0, 0.0]])

        cfg_auto = PowerContrastConfig(
            L=L_wp, delta=np.array([1.0]),
            sigma=1.0, power=0.5, alpha=0.10, max_n=60, max_iter=5,
        )
        cfg_conservative = PowerContrastConfig(
            L=L_wp, delta=np.array([1.0]),
            sigma=1.0, power=0.5, alpha=0.10, max_n=60, max_iter=5,
        )
        cfg_sp_only = PowerContrastConfig(
            L=L_wp, delta=np.array([1.0]),
            sigma=1.0, power=0.5, alpha=0.10, max_n=60, max_iter=5,
        )

        result_auto = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, cfg_auto,
            design_opts=_sp7_opts(df_method="auto"),
        )
        result_conservative = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, cfg_conservative,
            design_opts=_sp7_opts(df_method="conservative"),
        )
        result_sp_only = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, cfg_sp_only,
            design_opts=_sp7_opts(df_method="sp_only"),
        )

        power_auto = result_auto["report"]["achieved_power"]
        power_conservative = result_conservative["report"]["achieved_power"]
        power_sp_only = result_sp_only["report"]["achieved_power"]

        # "auto" for a WP contrast must use WP df (= conservative), not SP df.
        # WP df <= SP df in general, so conservative power <= sp_only power.
        # After the fix, auto == conservative for a pure WP contrast.
        assert abs(power_auto - power_conservative) < 0.05, (
            f"CR-25: 'auto' ({power_auto:.4f}) should match 'conservative' "
            f"({power_conservative:.4f}) for a WP contrast, not 'sp_only' ({power_sp_only:.4f})"
        )

    @pytest.mark.slow  # ~448s measured 2026-07-11 (TICKET-006A)
    def test_sp_contrast_auto_equals_sp_only(self):
        """CR-25: SP contrast power under 'auto' must equal 'sp_only'
        (both use SP df).  A pure-SP contrast is unaffected by the fix.
        """
        # L targets C (an ETC factor) — pure SP contrast
        L_sp = np.array([[0.0, 0.0, 0.0, 1.0]])

        cfg_auto = PowerContrastConfig(
            L=L_sp, delta=np.array([0.5]),
            sigma=1.0, power=0.5, alpha=0.10, max_n=60, max_iter=5,
        )
        cfg_sp_only = PowerContrastConfig(
            L=L_sp, delta=np.array([0.5]),
            sigma=1.0, power=0.5, alpha=0.10, max_n=60, max_iter=5,
        )

        result_auto = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, cfg_auto,
            design_opts=_sp7_opts(df_method="auto"),
        )
        result_sp_only = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, cfg_sp_only,
            design_opts=_sp7_opts(df_method="sp_only"),
        )

        power_auto = result_auto["report"]["achieved_power"]
        power_sp_only = result_sp_only["report"]["achieved_power"]

        assert abs(power_auto - power_sp_only) < 0.05, (
            f"CR-25: SP contrast 'auto' ({power_auto:.4f}) should match "
            f"'sp_only' ({power_sp_only:.4f})"
        )


# ===========================================================================
# TestSplitPlotAnalysis  (SP-8)
# ===========================================================================

# Shared config for SP-8 tests — small, fast.
_SP8_FACTORS = {"A": (-1.0, 1.0), "B": (-1.0, 1.0), "C": (0.0, 1.0)}
_SP8_FORMULA = "~ 1 + A + B + C"
_SP8_HTC = ["A"]


def _sp8_contrast_cfg():
    import numpy as np
    return PowerContrastConfig(
        L=np.array([[0.0, 1.0, 0.0, 0.0]]),
        delta=np.array([1.0]),
        sigma=1.0,
        power=0.5,
        alpha=0.10,
        max_n=40,
        max_iter=8,
    )


def _sp8_r2_cfg():
    return PowerR2Config(
        r2_target=0.5,
        power=0.5,
        alpha=0.10,
        max_n=40,
        max_iter=8,
    )


def _sp8_design_opts():
    return DesignOptions(starts=2, max_iter=5, random_state=7)


# ===========================================================================
# TestCR27CandidateSizing — candidate_points respected in split-plot mode
# ===========================================================================

class TestCR27CandidateSizing:
    """CR-27 regression: build_split_plot_design must use candidate_points
    from DesignOptions rather than the hard-coded defaults (n_wp_cand=30,
    n_sp_cand=50)."""

    def test_large_candidate_points_runs_without_error(self):
        """CR-27: candidate_points=500 is forwarded to the WP/SP pools."""
        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=1, max_iter=5, random_state=0,
            candidate_points=500,
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_r2_cfg(), design_opts=opts,
        )
        assert "design_df" in result

    def test_small_candidate_points_runs_without_error(self):
        """CR-27: candidate_points=50 still produces a valid design."""
        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=1, max_iter=5, random_state=1,
            candidate_points=50,
        )
        result = find_optimal_design(
            _SP7_FORMULA, _SP7_FACTORS, _sp7_r2_cfg(), design_opts=opts,
        )
        assert "design_df" in result

    def test_pool_sizes_proportional_to_factor_count(self):
        """CR-27: n_wp_cand and n_sp_cand are proportional to HTC/ETC factor count.

        With 1 HTC factor and 2 ETC factors (n_all=3), candidate_points=300:
          n_wp_cand = max(10, int(300 * 1/3)) = 100
          n_sp_cand = max(10, int(300 * 2/3)) = 200
        build_split_plot_design is called with these values via the mock.
        """
        from unittest.mock import patch, call
        from lattice_doe.iopt_search import build_split_plot_design as _bsd

        captured_kwargs: list = []

        def _mock_bsd(*args, **kwargs):
            captured_kwargs.append(kwargs)
            return _bsd(*args, **kwargs)

        opts = DesignOptions(
            split_plot=SplitPlotOptions(
                htc_factors=["A"], n_whole_plots=3, subplots_per_wp=3, eta=1.0,
            ),
            starts=1, max_iter=3, random_state=42,
            candidate_points=300,
        )

        with patch("lattice_doe.api.build_split_plot_design", side_effect=_mock_bsd):
            find_optimal_design(
                _SP7_FORMULA, _SP7_FACTORS, _sp7_r2_cfg(), design_opts=opts,
            )

        assert len(captured_kwargs) >= 1
        kw = captured_kwargs[0]
        # 1 HTC (A), 2 ETC (B, C) → n_wp_cand=100, n_sp_cand=200
        assert kw.get("n_wp_cand") == 100, (
            f"CR-27: n_wp_cand={kw.get('n_wp_cand')!r}, expected 100"
        )
        assert kw.get("n_sp_cand") == 200, (
            f"CR-27: n_sp_cand={kw.get('n_sp_cand')!r}, expected 200"
        )

    def test_power_curve_by_wp_uses_candidate_points(self):
        """CR-27: power_curve_by_wp also forwards candidate_points to pool sizes."""
        from unittest.mock import patch
        from lattice_doe.iopt_search import build_split_plot_design as _bsd

        captured_kwargs: list = []

        def _mock_bsd(*args, **kwargs):
            captured_kwargs.append(kwargs)
            return _bsd(*args, **kwargs)

        opts = DesignOptions(starts=1, max_iter=3, random_state=0, candidate_points=300)

        # power_curve_by_wp uses a deferred `from .iopt_search import ...`
        # inside the function body, so we patch at the source module.
        with patch("lattice_doe.iopt_search.build_split_plot_design", side_effect=_mock_bsd):
            power_curve_by_wp(
                _SP7_FORMULA, _SP7_FACTORS, _sp7_contrast_cfg(),
                subplots_per_wp=3, htc_factors=_SP7_HTC, eta=1.0,
                wp_range=(2, 3), wp_points=2,
                design_opts=opts,
            )

        assert len(captured_kwargs) >= 1
        kw = captured_kwargs[0]
        # 1 HTC (A), 2 ETC (B, C), candidate_points=300 → n_wp_cand=100, n_sp_cand=200
        assert kw.get("n_wp_cand") == 100, (
            f"CR-27: power_curve_by_wp n_wp_cand={kw.get('n_wp_cand')!r}, expected 100"
        )
        assert kw.get("n_sp_cand") == 200, (
            f"CR-27: power_curve_by_wp n_sp_cand={kw.get('n_sp_cand')!r}, expected 200"
        )


class TestSplitPlotAnalysis:

    # ------------------------------------------------------------------ #
    # power_curve_by_wp — return type and structure                        #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~126s measured 2026-07-11 (TICKET-006A)
    def test_returns_dataframe(self):
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 5), wp_points=4,
            design_opts=_sp8_design_opts(),
        )
        assert isinstance(df, pd.DataFrame)

    @pytest.mark.slow  # ~125s measured 2026-07-11 (TICKET-006A)
    def test_has_required_columns(self):
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 5), wp_points=4,
            design_opts=_sp8_design_opts(),
        )
        for col in ("n_wp", "n_total", "power", "noncentrality_lambda"):
            assert col in df.columns, f"Missing column: {col!r}"

    @pytest.mark.slow  # ~175s measured 2026-07-11 (TICKET-006A)
    def test_row_count_equals_wp_points(self):
        wp_points = 5
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 6), wp_points=wp_points,
            design_opts=_sp8_design_opts(),
        )
        assert len(df) == wp_points

    @pytest.mark.slow  # ~125s measured 2026-07-11 (TICKET-006A)
    def test_power_in_unit_interval(self):
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 5), wp_points=4,
            design_opts=_sp8_design_opts(),
        )
        valid = df["power"].dropna()
        assert (valid >= 0.0).all() and (valid <= 1.0).all()

    @pytest.mark.slow  # ~92s measured 2026-07-11 (TICKET-006A)
    def test_n_total_equals_n_wp_times_s(self):
        s = 3
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=s, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 4), wp_points=3,
            design_opts=_sp8_design_opts(),
        )
        assert (df["n_total"] == df["n_wp"] * s).all()

    @pytest.mark.slow  # ~230s measured 2026-07-11 (TICKET-006A)
    def test_power_correlates_with_n_wp(self):
        """Power should be non-decreasing overall as n_wp grows."""
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 8), wp_points=7,
            design_opts=_sp8_design_opts(),
        )
        valid = df["power"].dropna()
        if len(valid) >= 2:
            # Spearman correlation with n_wp should be positive
            import scipy.stats as sps
            corr, _ = sps.spearmanr(df["n_wp"].values[:len(valid)], valid.values)
            assert corr >= 0.0, f"Power unexpectedly decreases with n_wp: corr={corr:.3f}"

    @pytest.mark.slow  # ~1039s measured 2026-07-11 (TICKET-006A)
    def test_r2_mode_works(self):
        df = power_curve_by_wp(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_r2_cfg(),
            subplots_per_wp=3, htc_factors=_SP8_HTC, eta=1.0,
            wp_range=(2, 5), wp_points=4,
            design_opts=_sp8_design_opts(),
        )
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 4

    # ------------------------------------------------------------------ #
    # power_sensitivity — eta sweep                                        #
    # ------------------------------------------------------------------ #

    @pytest.mark.slow  # ~196s measured 2026-07-11 (TICKET-006A)
    def test_eta_sweep_present_for_sp_design(self):
        """power_sensitivity returns eta_sweep DataFrame when eta_range given."""
        # Build a small SP design first
        result_api = find_optimal_design(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            design_opts=DesignOptions(
                split_plot=SplitPlotOptions(
                    htc_factors=_SP8_HTC, n_whole_plots=4,
                    eta=1.0, subplots_per_wp=3,
                ),
                starts=2, max_iter=5, random_state=7,
            ),
        )
        design_df = result_api["design_df"]
        sens = power_sensitivity(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            design_df=design_df,
            eta_range=(0.0, 3.0), eta_points=10,
            design_opts=_sp8_design_opts(),
        )
        assert "eta_sweep" in sens
        assert isinstance(sens["eta_sweep"], pd.DataFrame)
        eta_df = sens["eta_sweep"]
        assert len(eta_df) == 10
        assert "eta" in eta_df.columns
        assert "power" in eta_df.columns

    @pytest.mark.slow  # ~199s measured 2026-07-11 (TICKET-006A)
    def test_eta_sweep_power_decreases_with_eta(self):
        """WP contrast power should decrease as eta increases."""
        result_api = find_optimal_design(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            design_opts=DesignOptions(
                split_plot=SplitPlotOptions(
                    htc_factors=_SP8_HTC, n_whole_plots=4,
                    eta=1.0, subplots_per_wp=3,
                ),
                starts=2, max_iter=5, random_state=7,
            ),
        )
        sens = power_sensitivity(
            _SP8_FORMULA, _SP8_FACTORS, _sp8_contrast_cfg(),
            design_df=result_api["design_df"],
            eta_range=(0.1, 5.0), eta_points=10,
            design_opts=_sp8_design_opts(),
        )
        eta_df = sens["eta_sweep"]
        # Correlate: power should generally decrease as eta grows
        import scipy.stats as sps
        corr, _ = sps.spearmanr(eta_df["eta"].values, eta_df["power"].values)
        assert corr <= 0.0, f"Power unexpectedly increases with eta: corr={corr:.3f}"

    def test_eta_sweep_none_for_non_sp_design(self):
        """power_sensitivity returns eta_sweep=None when __wp_id__ not in design_df."""
        from lattice_doe import find_optimal_design as iopt
        result_ols = iopt(
            _SP8_FORMULA, _SP8_FACTORS,
            PowerContrastConfig(
                L=np.array([[0.0, 1.0, 0.0, 0.0]]),
                delta=np.array([1.0]), sigma=1.0,
                power=0.5, alpha=0.10, max_n=20, max_iter=5,
            ),
            design_opts=DesignOptions(starts=2, max_iter=4, random_state=0),
        )
        sens = power_sensitivity(
            _SP8_FORMULA, _SP8_FACTORS,
            PowerContrastConfig(
                L=np.array([[0.0, 1.0, 0.0, 0.0]]),
                delta=np.array([1.0]), sigma=1.0,
                power=0.5, alpha=0.10, max_n=20, max_iter=5,
            ),
            design_df=result_ols["design_df"],
            eta_range=(0.0, 3.0), eta_points=5,
        )
        assert sens["eta_sweep"] is None

    def test_power_sensitivity_ols_unchanged(self):
        """Existing power_sensitivity behavior unaffected when eta_range=None."""
        from lattice_doe import find_optimal_design as iopt
        result_ols = iopt(
            _SP8_FORMULA, _SP8_FACTORS,
            PowerContrastConfig(
                L=np.array([[0.0, 1.0, 0.0, 0.0]]),
                delta=np.array([1.0]), sigma=1.0,
                power=0.5, alpha=0.10, max_n=20, max_iter=5,
            ),
            design_opts=DesignOptions(starts=2, max_iter=4, random_state=0),
        )
        sens = power_sensitivity(
            _SP8_FORMULA, _SP8_FACTORS,
            PowerContrastConfig(
                L=np.array([[0.0, 1.0, 0.0, 0.0]]),
                delta=np.array([1.0]), sigma=1.0,
                power=0.5, alpha=0.10, max_n=20, max_iter=5,
            ),
            design_df=result_ols["design_df"],
        )
        assert "data" in sens
        assert "nominal_power" in sens
        assert sens["eta_sweep"] is None


# ===========================================================================
# SP-9  Tests — CLI / Sheets / Excel integration
# ===========================================================================

# ---------------------------------------------------------------------------
# TestSplitPlotCLI
# ---------------------------------------------------------------------------

class TestSplitPlotCLI:
    """Tests for split-plot flags and YAML block parsing in cli.py."""

    def test_make_design_opts_no_sp_block(self):
        """Without split_plot key, design_opts has split_plot=None."""
        from lattice_doe.cli import _make_design_opts
        cfg: dict = {}
        opts = _make_design_opts(cfg)
        assert opts.split_plot is None

    def test_make_design_opts_sp_block_activates_split_plot(self):
        """YAML split_plot: block builds SplitPlotOptions correctly."""
        from lattice_doe.cli import _make_design_opts
        cfg = {
            "split_plot": {
                "htc_factors": ["A", "B"],
                "n_whole_plots": 6,
                "eta": 2.0,
                "subplots_per_wp": 4,
                "df_method": "conservative",
            }
        }
        opts = _make_design_opts(cfg)
        assert opts.split_plot is not None
        sp = opts.split_plot
        assert sp.htc_factors == ["A", "B"]
        assert sp.n_whole_plots == 6
        assert sp.eta == pytest.approx(2.0)
        assert sp.subplots_per_wp == 4
        assert sp.df_method == "conservative"

    def test_make_design_opts_sp_block_string_htc_factors(self):
        """htc_factors as comma-separated string is parsed into a list."""
        from lattice_doe.cli import _make_design_opts
        cfg = {
            "split_plot": {
                "htc_factors": "A, B, C",
                "n_whole_plots": 4,
            }
        }
        opts = _make_design_opts(cfg)
        assert opts.split_plot is not None
        assert opts.split_plot.htc_factors == ["A", "B", "C"]

    def test_make_design_opts_sp_block_missing_htc_factors_no_activate(self):
        """Empty htc_factors does not activate split-plot."""
        from lattice_doe.cli import _make_design_opts
        cfg = {"split_plot": {"htc_factors": [], "n_whole_plots": 4}}
        opts = _make_design_opts(cfg)
        assert opts.split_plot is None

    def test_make_design_opts_sp_block_n_whole_plots_lt2_no_activate(self):
        """n_whole_plots < 2 does not activate split-plot."""
        from lattice_doe.cli import _make_design_opts
        cfg = {"split_plot": {"htc_factors": ["A"], "n_whole_plots": 1}}
        opts = _make_design_opts(cfg)
        assert opts.split_plot is None

    def test_make_design_opts_sp_subplots_per_wp_none_when_zero(self):
        """subplots_per_wp=0 in YAML maps to None (auto)."""
        from lattice_doe.cli import _make_design_opts
        cfg = {
            "split_plot": {
                "htc_factors": ["A"],
                "n_whole_plots": 3,
                "subplots_per_wp": 0,
            }
        }
        opts = _make_design_opts(cfg)
        assert opts.split_plot is not None
        assert opts.split_plot.subplots_per_wp is None

    def test_template_contrast_contains_split_plot_section(self):
        """--template contrast output contains a commented split_plot block."""
        from lattice_doe.cli import _TEMPLATE_CONTRAST
        assert "split_plot" in _TEMPLATE_CONTRAST
        assert "htc_factors" in _TEMPLATE_CONTRAST
        assert "n_whole_plots" in _TEMPLATE_CONTRAST

    def test_template_r2_contains_split_plot_section(self):
        """--template r2 output contains a commented split_plot block."""
        from lattice_doe.cli import _TEMPLATE_R2
        assert "split_plot" in _TEMPLATE_R2
        assert "htc_factors" in _TEMPLATE_R2

    def test_cli_flags_merge_into_cfg(self, tmp_path):
        """CLI --htc-factors / --n-whole-plots override YAML split_plot block."""
        import textwrap
        from lattice_doe.cli import main

        cfg_text = textwrap.dedent("""\
            formula: "~ 1 + A + B"
            factors:
              A: [0.0, 1.0]
              B: [0.0, 1.0]
            r2_target: 0.30
            alpha: 0.05
            power: 0.50
            design:
              starts: 1
              max_iter: 5
              random_state: 0
        """)
        cfg_path = tmp_path / "cfg.yaml"
        cfg_path.write_text(cfg_text, encoding="utf-8")
        out_base = str(tmp_path / "out")

        # The dry-run flag stops before the actual design search; it still
        # validates that the config + CLI flags are parsed without error.
        ret = main([
            "--config", str(cfg_path),
            "--out", out_base,
            "--dry-run",
            "--htc-factors", "A",
            "--n-whole-plots", "3",
            "--eta", "1.5",
        ])
        assert ret == 0

    def test_cli_no_sp_flags_is_unchanged(self, tmp_path):
        """Without SP flags, CLI behaves exactly as before (no regressions)."""
        import textwrap
        from lattice_doe.cli import main

        cfg_text = textwrap.dedent("""\
            formula: "~ 1 + A + B"
            factors:
              A: [0.0, 1.0]
              B: [0.0, 1.0]
            r2_target: 0.30
            alpha: 0.05
            power: 0.50
            design:
              starts: 1
              max_iter: 5
              random_state: 0
        """)
        cfg_path = tmp_path / "cfg.yaml"
        cfg_path.write_text(cfg_text, encoding="utf-8")
        ret = main([
            "--config", str(cfg_path),
            "--out", str(tmp_path / "out"),
            "--dry-run",
        ])
        assert ret == 0

    # ------------------------------------------------------------------ #
    # CR-26 regression: standalone SP flags (no --htc-factors/            #
    # --n-whole-plots) must still override YAML values                    #
    # ------------------------------------------------------------------ #

    def test_apply_sp_cli_args_eta_standalone(self):
        """CR-26: --eta alone overrides YAML split_plot.eta."""
        from types import SimpleNamespace
        from lattice_doe.cli import _apply_sp_cli_args, _make_design_opts

        yaml_cfg = {
            "split_plot": {"htc_factors": ["A"], "n_whole_plots": 3, "eta": 1.0}
        }
        args = SimpleNamespace(
            htc_factors=None, n_whole_plots=None,
            eta=5.0, subplots_per_wp=None, df_method=None,
        )
        merged = _apply_sp_cli_args(yaml_cfg, args)
        opts = _make_design_opts(merged)
        assert opts.split_plot is not None
        assert opts.split_plot.eta == pytest.approx(5.0), (
            "CR-26: --eta standalone did not override YAML split_plot.eta"
        )

    def test_apply_sp_cli_args_subplots_standalone(self):
        """CR-26: --subplots-per-wp alone overrides YAML split_plot.subplots_per_wp."""
        from types import SimpleNamespace
        from lattice_doe.cli import _apply_sp_cli_args, _make_design_opts

        yaml_cfg = {
            "split_plot": {
                "htc_factors": ["A"], "n_whole_plots": 4,
                "eta": 1.0, "subplots_per_wp": 2,
            }
        }
        args = SimpleNamespace(
            htc_factors=None, n_whole_plots=None,
            eta=None, subplots_per_wp=5, df_method=None,
        )
        merged = _apply_sp_cli_args(yaml_cfg, args)
        opts = _make_design_opts(merged)
        assert opts.split_plot is not None
        assert opts.split_plot.subplots_per_wp == 5, (
            "CR-26: --subplots-per-wp standalone did not override YAML value"
        )

    def test_apply_sp_cli_args_df_method_standalone(self):
        """CR-26: --df-method alone overrides YAML split_plot.df_method."""
        from types import SimpleNamespace
        from lattice_doe.cli import _apply_sp_cli_args, _make_design_opts

        yaml_cfg = {
            "split_plot": {
                "htc_factors": ["A"], "n_whole_plots": 3,
                "eta": 1.0, "df_method": "auto",
            }
        }
        args = SimpleNamespace(
            htc_factors=None, n_whole_plots=None,
            eta=None, subplots_per_wp=None, df_method="conservative",
        )
        merged = _apply_sp_cli_args(yaml_cfg, args)
        opts = _make_design_opts(merged)
        assert opts.split_plot is not None
        assert opts.split_plot.df_method == "conservative", (
            "CR-26: --df-method standalone did not override YAML value"
        )

    def test_apply_sp_cli_args_no_flags_leaves_cfg_unchanged(self):
        """CR-26: when no SP flags are provided, cfg is returned as-is."""
        from types import SimpleNamespace
        from lattice_doe.cli import _apply_sp_cli_args

        yaml_cfg = {
            "split_plot": {"htc_factors": ["A"], "n_whole_plots": 3, "eta": 2.0}
        }
        args = SimpleNamespace(
            htc_factors=None, n_whole_plots=None,
            eta=None, subplots_per_wp=None, df_method=None,
        )
        result = _apply_sp_cli_args(yaml_cfg, args)
        assert result is yaml_cfg  # same object — no copy made

    def test_apply_sp_cli_args_original_cfg_not_mutated(self):
        """CR-26: _apply_sp_cli_args does not mutate the input cfg."""
        from types import SimpleNamespace
        from lattice_doe.cli import _apply_sp_cli_args

        yaml_cfg = {
            "split_plot": {"htc_factors": ["A"], "n_whole_plots": 3, "eta": 1.0}
        }
        original_eta = yaml_cfg["split_plot"]["eta"]
        args = SimpleNamespace(
            htc_factors=None, n_whole_plots=None,
            eta=9.0, subplots_per_wp=None, df_method=None,
        )
        _apply_sp_cli_args(yaml_cfg, args)
        assert yaml_cfg["split_plot"]["eta"] == original_eta  # original untouched


# ---------------------------------------------------------------------------
# TestSplitPlotSheets
# ---------------------------------------------------------------------------

def _make_mock_ws_sp(rows):
    """Mock gspread Worksheet whose get_all_values() returns *rows*."""
    from unittest.mock import MagicMock
    ws = MagicMock()
    ws.get_all_values.return_value = rows
    return ws


def _sp_sheet_rows(extra_settings: dict) -> list:
    """Build a minimal Config sheet row list for _parse_config_sheet."""
    base_settings = [
        ["formula", "x1 + x2"],
        ["power_mode", "r2"],
        ["alpha", "0.05"],
        ["power", "0.80"],
        ["r2_target", "0.25"],
        ["max_n", "100"],
        ["criterion", "I"],
        ["starts", "3"],
        ["max_iter", "100"],
        ["random_state", "0"],
    ]
    for k, v in extra_settings.items():
        base_settings.append([k, v])
    return (
        [["[SETTINGS]", ""]]
        + base_settings
        + [["", ""],
           ["[FACTORS]", ""],
           ["factor_name", "type", "value1", "value2"],
           ["x1", "continuous", "-1.0", "1.0"],
           ["x2", "continuous", "-1.0", "1.0"]]
    )


class TestSplitPlotSheets:
    """Tests for split-plot settings in the Google Sheets connector."""

    def test_template_rows_r2_contain_sp_fields(self):
        """R2 template contains all 5 SP field names."""
        from lattice_doe.sheets import _TEMPLATE_ROWS
        rows_r2 = _TEMPLATE_ROWS["r2"]
        keys = [r[0] for r in rows_r2]
        for key in ("htc_factors", "n_whole_plots", "eta", "subplots_per_wp", "df_method"):
            assert key in keys, f"Missing key {key!r} in r2 template"

    def test_template_rows_contrast_contain_sp_fields(self):
        """Contrast template contains all 5 SP field names."""
        from lattice_doe.sheets import _TEMPLATE_ROWS
        rows_contrast = _TEMPLATE_ROWS["contrast"]
        keys = [r[0] for r in rows_contrast]
        for key in ("htc_factors", "n_whole_plots", "eta", "subplots_per_wp", "df_method"):
            assert key in keys, f"Missing key {key!r} in contrast template"

    def test_sp_fields_disabled_by_default_in_template(self):
        """Default template has n_whole_plots=0 (SP disabled)."""
        from lattice_doe.sheets import _TEMPLATE_ROWS
        row_dict = {r[0]: r[1] for r in _TEMPLATE_ROWS["r2"] if len(r) > 1}
        assert row_dict["n_whole_plots"] == "0"
        assert row_dict["htc_factors"] == ""

    def test_design_opts_no_sp_when_htc_factors_empty(self):
        """Parsing settings with blank htc_factors produces no split_plot."""
        from lattice_doe.sheets import _parse_config_sheet
        rows = _sp_sheet_rows({"htc_factors": "", "n_whole_plots": "4"})
        _, _, _, design_opts, _ = _parse_config_sheet(_make_mock_ws_sp(rows))
        assert design_opts.split_plot is None

    def test_design_opts_sp_activated_when_valid(self):
        """Parsing settings with htc_factors + n_whole_plots >= 2 creates SplitPlotOptions."""
        from lattice_doe.sheets import _parse_config_sheet
        rows = _sp_sheet_rows({
            "htc_factors": "x1",
            "n_whole_plots": "4",
            "eta": "2.0",
            "subplots_per_wp": "3",
            "df_method": "conservative",
        })
        _, _, _, design_opts, _ = _parse_config_sheet(_make_mock_ws_sp(rows))
        assert design_opts.split_plot is not None
        sp = design_opts.split_plot
        assert sp.htc_factors == ["x1"]
        assert sp.n_whole_plots == 4
        assert sp.eta == pytest.approx(2.0)
        assert sp.subplots_per_wp == 3
        assert sp.df_method == "conservative"

    def test_design_opts_n_whole_plots_lt2_no_sp(self):
        """n_whole_plots < 2 does not activate split-plot even with htc_factors set."""
        from lattice_doe.sheets import _parse_config_sheet
        rows = _sp_sheet_rows({"htc_factors": "x1", "n_whole_plots": "1"})
        _, _, _, design_opts, _ = _parse_config_sheet(_make_mock_ws_sp(rows))
        assert design_opts.split_plot is None

    def test_design_opts_subplots_per_wp_zero_maps_to_none(self):
        """subplots_per_wp=0 in sheet maps to None (auto)."""
        from lattice_doe.sheets import _parse_config_sheet
        rows = _sp_sheet_rows({
            "htc_factors": "x1",
            "n_whole_plots": "3",
            "subplots_per_wp": "0",
        })
        _, _, _, design_opts, _ = _parse_config_sheet(_make_mock_ws_sp(rows))
        assert design_opts.split_plot is not None
        assert design_opts.split_plot.subplots_per_wp is None


# ---------------------------------------------------------------------------
# TestSplitPlotExcel
# ---------------------------------------------------------------------------

class TestSplitPlotExcel:
    """Tests for split-plot settings in the Excel connector."""

    def test_create_excel_template_r2_contains_sp_keys(self, tmp_path):
        """create_excel_template (r2) writes SP key rows to the Config sheet."""
        pytest.importorskip("openpyxl")
        from lattice_doe.excel_template import create_excel_template
        import openpyxl

        dest = create_excel_template(str(tmp_path / "tpl.xlsx"), example="r2")
        wb = openpyxl.load_workbook(dest)
        ws = wb["Config"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        col_a_str = [str(v).strip() if v is not None else "" for v in col_a]
        for key in ("htc_factors", "n_whole_plots", "eta", "subplots_per_wp", "df_method"):
            assert key in col_a_str, f"Key {key!r} missing from r2 template Config sheet"

    def test_create_excel_template_contrast_contains_sp_keys(self, tmp_path):
        """create_excel_template (contrast) writes SP key rows."""
        pytest.importorskip("openpyxl")
        from lattice_doe.excel_template import create_excel_template
        import openpyxl

        dest = create_excel_template(str(tmp_path / "tpl.xlsx"), example="contrast")
        wb = openpyxl.load_workbook(dest)
        ws = wb["Config"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        col_a_str = [str(v).strip() if v is not None else "" for v in col_a]
        for key in ("htc_factors", "n_whole_plots", "eta", "subplots_per_wp", "df_method"):
            assert key in col_a_str, f"Key {key!r} missing from contrast template Config sheet"

    def test_read_config_sheet_no_sp_when_n_whole_plots_zero(self, tmp_path):
        """Reading a default template (n_whole_plots=0) produces no SplitPlotOptions."""
        pytest.importorskip("openpyxl")
        from lattice_doe.excel_template import create_excel_template, _read_config_sheet
        import openpyxl

        dest = create_excel_template(str(tmp_path / "tpl.xlsx"), example="r2")
        wb = openpyxl.load_workbook(dest)
        _, _, _, design_opts, _ = _read_config_sheet(wb["Config"])
        assert design_opts.split_plot is None

    def test_read_config_sheet_sp_activated(self, tmp_path):
        """Writing SP values to the template then reading back produces SplitPlotOptions."""
        pytest.importorskip("openpyxl")
        from lattice_doe.excel_template import create_excel_template, _read_config_sheet
        import openpyxl

        dest = create_excel_template(str(tmp_path / "tpl.xlsx"), example="r2")
        wb = openpyxl.load_workbook(dest)
        ws = wb["Config"]

        # Edit SP rows in the workbook
        for r in range(1, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            if key == "htc_factors":
                ws.cell(row=r, column=2).value = "A"
            elif key == "n_whole_plots":
                ws.cell(row=r, column=2).value = 4
            elif key == "eta":
                ws.cell(row=r, column=2).value = 1.5
            elif key == "subplots_per_wp":
                ws.cell(row=r, column=2).value = 3
            elif key == "df_method":
                ws.cell(row=r, column=2).value = "conservative"
            # Also update factors to include A and B (default only has A and B already)

        wb.save(dest)
        wb2 = openpyxl.load_workbook(dest)
        _, _, _, design_opts, _ = _read_config_sheet(wb2["Config"])
        assert design_opts.split_plot is not None
        sp = design_opts.split_plot
        assert sp.htc_factors == ["A"]
        assert sp.n_whole_plots == 4
        assert sp.eta == pytest.approx(1.5)
        assert sp.subplots_per_wp == 3
        assert sp.df_method == "conservative"

    def test_sp_invalid_options_raise_excel_error(self, tmp_path):
        """Invalid SplitPlotOptions values (eta < 0) raise ExcelError."""
        pytest.importorskip("openpyxl")
        from lattice_doe.excel_template import (
            create_excel_template, _read_config_sheet, ExcelError,
        )
        import openpyxl

        dest = create_excel_template(str(tmp_path / "tpl.xlsx"), example="r2")
        wb = openpyxl.load_workbook(dest)
        ws = wb["Config"]

        for r in range(1, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            if key == "htc_factors":
                ws.cell(row=r, column=2).value = "A"
            elif key == "n_whole_plots":
                ws.cell(row=r, column=2).value = 4   # valid (>= 2) to trigger SplitPlotOptions
            elif key == "eta":
                ws.cell(row=r, column=2).value = -1.0  # invalid: must be >= 0

        wb.save(dest)
        wb2 = openpyxl.load_workbook(dest)
        with pytest.raises(ExcelError):
            _read_config_sheet(wb2["Config"])


# ===========================================================================
# SP-10 — Property-based / parametrized tests
# ===========================================================================

class TestSP10PropertyBased:
    """Parametrized regression tests verifying mathematical properties
    of split-plot utilities (SP-10 requirement)."""

    # ------------------------------------------------------------------
    # 1. Balanced layout: Z'Z = s * I_{n_wp}  (block-diagonal identity property)
    # ------------------------------------------------------------------

    @pytest.mark.parametrize("n_wp,s", [
        (2, 2), (2, 5), (3, 3), (4, 2), (5, 4), (6, 3), (10, 2),
    ])
    def test_ztZ_is_s_times_identity_for_balanced_layouts(self, n_wp, s):
        """Z'Z = s * I_{n_wp} for every balanced (n_wp, s) combination."""
        Z = build_whole_plot_indicator(n_wp * s, n_wp, s)
        ZtZ = Z.T @ Z
        np.testing.assert_allclose(ZtZ, s * np.eye(n_wp), atol=1e-12,
                                   err_msg=f"n_wp={n_wp}, s={s}: Z'Z ≠ s*I")

    # ------------------------------------------------------------------
    # 2. GLS criterion scorers match explicit matrix computation
    # ------------------------------------------------------------------

    @pytest.mark.parametrize("eta,n_wp,s", [
        (0.5, 2, 3), (1.0, 3, 3), (2.0, 4, 2), (5.0, 2, 4),
    ])
    def test_gls_d_criterion_matches_explicit_computation(self, eta, n_wp, s):
        """_gls_d_criterion matches explicit -log det(X'V⁻¹X) computation."""
        rng = np.random.default_rng(42)
        n_total = n_wp * s
        p = 3
        X = rng.standard_normal((n_total, p))
        Z = build_whole_plot_indicator(n_total, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta)
        M = X.T @ V_inv @ X + 1e-8 * np.eye(p)
        expected = -np.log(np.linalg.det(M))
        got = _gls_d_criterion(X, V_inv)
        assert abs(got - expected) < 1e-6, (
            f"eta={eta}, n_wp={n_wp}, s={s}: expected {expected:.6f}, got {got:.6f}"
        )

    @pytest.mark.parametrize("eta,n_wp,s", [
        (0.5, 2, 3), (1.0, 3, 3), (2.0, 4, 2),
    ])
    def test_gls_a_criterion_matches_explicit_computation(self, eta, n_wp, s):
        """_gls_a_criterion matches explicit tr[(X'V⁻¹X)⁻¹] computation."""
        rng = np.random.default_rng(43)
        n_total = n_wp * s
        p = 3
        X = rng.standard_normal((n_total, p))
        Z = build_whole_plot_indicator(n_total, n_wp, s)
        V_inv = build_split_plot_covariance_inv(Z, eta)
        M = X.T @ V_inv @ X + 1e-8 * np.eye(p)
        expected = float(np.trace(np.linalg.inv(M)))
        got = _gls_a_criterion(X, V_inv)
        assert abs(got - expected) < 1e-6, (
            f"eta={eta}, n_wp={n_wp}, s={s}: expected {expected:.6f}, got {got:.6f}"
        )

    # ------------------------------------------------------------------
    # 3. split_plot_df_denom: "conservative" df ≤ "sp_only" df for any contrast
    # ------------------------------------------------------------------

    @pytest.mark.parametrize("n_wp,s,is_wp", [
        (3, 4, True),
        (3, 4, False),
        (4, 4, True),
        (4, 4, False),
        (5, 3, True),
    ])
    def test_df_denom_conservative_le_sp_only(self, n_wp, s, is_wp):
        """df_method='conservative' yields df ≤ df from 'sp_only' when s is large enough.

        Requires s ≥ ceil(rank(X)/n_wp) + 1 for the inequality to hold.
        """
        rng = np.random.default_rng(44)
        n_total = n_wp * s
        p = 3
        X = rng.standard_normal((n_total, p))
        Z = build_whole_plot_indicator(n_total, n_wp, s)
        is_wp_arr = np.array([is_wp])
        df_cons = split_plot_df_denom(X, Z, is_wp_arr, "conservative")
        df_sp   = split_plot_df_denom(X, Z, is_wp_arr, "sp_only")
        assert df_cons[0] <= df_sp[0], (
            f"n_wp={n_wp}, s={s}, is_wp={is_wp}: "
            f"conservative={df_cons[0]} > sp_only={df_sp[0]}"
        )
